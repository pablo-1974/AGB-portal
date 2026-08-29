from __future__ import annotations



import io

import unicodedata

from utils.local_deps import ensure_local_deps

ensure_local_deps()
try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile

from fastapi.responses import HTMLResponse, RedirectResponse, Response



from auth import load_user_dep

from context import ctx

from db.groups import group_exists, list_groups_with_course, upsert_group_name

from utils.enums import PERM_GESTION_GRUPOS

from utils.permissions import has_permission



router = APIRouter(prefix="/admin/groups", tags=["admin_groups"])





def _templates(request: Request):

    return request.app.state.templates





def _require_groups_perm(user: dict) -> None:

    if not has_permission(user, PERM_GESTION_GRUPOS):

        raise HTTPException(status_code=403)





def _norm_header(value: object) -> str:

    text = str(value or "").strip().lower()

    text = "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")

    return " ".join(text.split())





@router.get("/", response_class=HTMLResponse)

def admin_groups(request: Request, user: dict = Depends(load_user_dep)):

    _require_groups_perm(user)

    return _templates(request).TemplateResponse(

        "admin/groups.html",

        ctx(

            request,

            user=user,

            title="Gestión de grupos",

            groups=list_groups_with_course(),

        ),

    )





@router.post("/import")

def admin_groups_import(

    user: dict = Depends(load_user_dep),

    file: UploadFile = File(...),

):

    _require_groups_perm(user)



    if not file.filename or not file.filename.lower().endswith(".xlsx"):

        return RedirectResponse("/admin/groups?status=error", status_code=303)



    try:

        wb = openpyxl.load_workbook(file.file)

        ws = wb.active

    except Exception:

        return RedirectResponse("/admin/groups?status=error", status_code=303)



    headers = [cell.value for cell in ws[1]]

    normalized = [_norm_header(h) for h in headers]

    idx = {name: pos for pos, name in enumerate(normalized) if name}



    name_idx = next((idx[k] for k in ("nombre", "grupo", "name") if k in idx), None)

    curso_idx = next((idx[k] for k in ("curso", "course", "nivel") if k in idx), None)

    if name_idx is None or curso_idx is None:

        return RedirectResponse("/admin/groups?status=error", status_code=303)



    created = 0

    updated = 0

    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):

        if not row:

            continue

        raw_name = row[name_idx] if name_idx < len(row) else None

        raw_curso = row[curso_idx] if curso_idx < len(row) else None

        if raw_name is None or not str(raw_name).strip():

            continue

        if raw_curso is None or not str(raw_curso).strip():

            skipped += 1

            continue

        try:

            existed_before = group_exists(str(raw_name))

            if upsert_group_name(str(raw_name), str(raw_curso)):

                created += 1

            else:

                if existed_before:

                    updated += 1

                else:

                    skipped += 1

        except ValueError:

            skipped += 1



    return RedirectResponse(

        f"/admin/groups?status=imported&created={created}&updated={updated}&skipped={skipped}",

        status_code=303,

    )





@router.get("/export")

def admin_groups_export(user: dict = Depends(load_user_dep)):

    _require_groups_perm(user)



    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Grupos"

    ws.append(["Nombre", "Curso"])

    for g in list_groups_with_course():

        ws.append([g["name"], g.get("curso") or ""])



    stream = io.BytesIO()

    wb.save(stream)

    stream.seek(0)



    return Response(

        stream.read(),

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={

            "Content-Disposition": "attachment; filename=grupos.xlsx",

        },

    )

