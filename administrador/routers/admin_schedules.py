from __future__ import annotations

import json
import logging
import tempfile
import unicodedata
from datetime import date, datetime
from urllib.parse import parse_qsl, urlencode

from utils.local_deps import ensure_local_deps

ensure_local_deps()
try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse

from auth import load_user_dep
from context import ctx
from ausencias.db import (
    add_action_log,
    apply_teacher_schedule_grid_edits,
    clear_schedule_cell,
    list_schedule_slots,
    list_teachers_for_schedule_selector,
    list_teachers_min,
    replace_schedule_slot,
    upsert_teacher_from_import,
)
from ausencias.services.pdf_schedule import generate_schedule_pdf
from db.users import get_user_by_email, get_user_by_id
from db.enrolled_subjects import map_materias_horario_por_grupo
from db.groups import list_groups
from utils.enums import PERM_GESTION_HORARIOS
from utils.permissions import has_permission
from utils.pdf_http import pdf_attachment_response, safe_pdf_filename
from utils.text import normalize_for_sort

router = APIRouter(prefix="/admin/schedules", tags=["admin_schedules"])

_log = logging.getLogger(__name__)


def _schedule_form_text(form, key: str) -> str:
    """Valor de campo de formulario como texto (evita ``.strip()`` sobre UploadFile)."""
    v = form.get(key)
    if v is None:
        return ""
    if hasattr(v, "read"):
        return ""
    return str(v).strip()


def _parse_urlencoded_body_flat(body: bytes) -> dict[str, str]:
    """Parsea ``application/x-www-form-urlencoded``; último valor gana por clave (como los navegadores)."""
    if not body:
        return {}
    text: str | None = None
    for enc in ("utf-8", "latin-1"):
        try:
            text = body.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = body.decode("utf-8", errors="replace")
    out: dict[str, str] = {}
    for k, v in parse_qsl(text, keep_blank_values=True):
        out[str(k)] = str(v).strip()
    return out


def _cell_kind_from_flat(flat: dict[str, str], prefix: str) -> str:
    """Tipo de celda; si el desplegable sigue en «–» pero hay datos, infiere CLASS, GUARD u OTHER."""
    raw_type = (flat.get(f"{prefix}type") or "NONE").upper()
    if raw_type not in {"NONE", "CLASS", "GUARD", "OTHER"}:
        raw_type = "NONE"
    group = (flat.get(f"{prefix}group") or "").strip()
    subject = (flat.get(f"{prefix}subject") or "").strip()
    guard_type = (flat.get(f"{prefix}guard_type") or "").strip()
    other_label = (flat.get(f"{prefix}other_label") or "").strip()
    if raw_type == "NONE":
        if group or subject:
            return "CLASS"
        if guard_type:
            return "GUARD"
        if other_label:
            return "OTHER"
        return "NONE"
    if raw_type == "OTHER":
        return "OTHER"
    if raw_type == "GUARD" and not guard_type and (group or subject):
        return "CLASS"
    if raw_type == "CLASS" and not group and not subject and guard_type:
        return "GUARD"
    if raw_type == "CLASS" and not group and not subject and other_label:
        return "OTHER"
    return raw_type


def _is_recreo_guard_type(guard_type: str | None) -> bool:
    return str(guard_type or "").strip().upper().startswith("G RECREO")


def _build_schedule_cells_from_flat(flat: dict[str, str]) -> list[dict]:
    """Cuadrícula 7×5 a partir de claves ``{h}_{d}_type`` etc."""
    cells: list[dict] = []
    for hour in range(7):
        for day in range(5):
            prefix = f"{hour}_{day}_"
            kind = _cell_kind_from_flat(flat, prefix)
            base = {"hour_index": hour, "day_index": day}
            if hour == RECREO_HOUR_INDEX:
                gt = (flat.get(f"{prefix}guard_type") or "").strip()
                if kind == "GUARD" and _is_recreo_guard_type(gt):
                    cells.append({**base, "kind": "GUARD", "guard_type": gt})
                else:
                    cells.append({**base, "kind": "NONE"})
                continue
            if kind == "NONE":
                cells.append({**base, "kind": "NONE"})
            elif kind == "CLASS":
                cells.append(
                    {
                        **base,
                        "kind": "CLASS",
                        "group": flat.get(f"{prefix}group", ""),
                        "room": flat.get(f"{prefix}room", ""),
                        "subject": flat.get(f"{prefix}subject", ""),
                    }
                )
            elif kind == "OTHER":
                cells.append(
                    {
                        **base,
                        "kind": "OTHER",
                        "subject": flat.get(f"{prefix}other_label", ""),
                    }
                )
            else:
                gt = (flat.get(f"{prefix}guard_type") or "").strip()
                if _is_recreo_guard_type(gt):
                    cells.append({**base, "kind": "NONE"})
                else:
                    cells.append(
                        {
                            **base,
                            "kind": "GUARD",
                            "guard_type": gt,
                        }
                    )
    return cells


def _flat_from_starlette_form(form) -> dict[str, str]:
    flat: dict[str, str] = {}
    for key, val in form.multi_items():
        if hasattr(val, "read"):
            continue
        flat[str(key)] = str(val).strip()
    return flat


def _schedule_form_has_grid(flat: dict[str, str]) -> bool:
    """Al menos una celda ``{h}_{d}_type`` (cuadrícula 7×5)."""
    return any(k.endswith("_type") for k in flat)


async def _read_schedule_post_flat(request: Request) -> dict[str, str]:
    """Lee el POST del horario (urlencoded estándar del navegador o multipart)."""
    ct = (request.headers.get("content-type") or "").lower()
    if "multipart/form-data" in ct:
        try:
            return _flat_from_starlette_form(await request.form())
        except Exception:
            _log.warning("request.form() falló (multipart horario)", exc_info=True)
            return {}

    body = await request.body()
    flat = _parse_urlencoded_body_flat(body)
    if _schedule_form_has_grid(flat):
        return flat
    if not flat:
        try:
            return _flat_from_starlette_form(await request.form())
        except Exception:
            _log.warning("request.form() falló (urlencoded horario)", exc_info=True)
    return flat


HOUR_LABELS = ("1ª", "2ª", "3ª", "Recreo", "4ª", "5ª", "6ª")
RECREO_HOUR_INDEX = 3
HOURS = {
    "1ª": 0,
    "2ª": 1,
    "3ª": 2,
    "recreo": 3,
    "4ª": 4,
    "5ª": 5,
    "6ª": 6,
    "1a": 0,
    "2a": 1,
    "3a": 2,
    "4a": 4,
    "5a": 5,
    "6a": 6,
}

DAYS = {"lunes": 0, "martes": 1, "miercoles": 2, "jueves": 3, "viernes": 4}

_DAY_ALIASES = {
    "lun": 0,
    "mar": 1,
    "mier": 2,
    "mie": 2,
    "jue": 3,
    "vie": 4,
    "monday": 0,
    "tuesday": 1,
    "wednesday": 2,
    "thursday": 3,
    "friday": 4,
    "mon": 0,
    "tue": 1,
    "wed": 2,
    "thu": 3,
    "fri": 4,
}

GUARD_LABELS = frozenset({"G AULA", "G RECREO PATIO", "G RECREO PASILLO"})


def _templates(request: Request):
    return request.app.state.templates


def _require_permission(user: dict) -> None:
    if not has_permission(user, PERM_GESTION_HORARIOS):
        raise HTTPException(status_code=403)


def _norm_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")
    return " ".join(text.split())


def _hour_cell_index(value: object) -> int | None:
    """
    Fila en la cuadrícula 0..6: 1ª, 2ª, 3ª, Recreo, 4ª, 5ª, 6ª.

    Muchas plantillas guardan la hora como número entero 1..7 (orden de la cuadrícula,
    recreo = 4); openpyxl puede devolver ``int``/``float`` o cadenas tipo ``"4.0"``.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            n = int(round(float(value)))
        except (TypeError, ValueError):
            return None
        if 1 <= n <= 7:
            return n - 1
        return None

    raw = str(value).strip()
    if not raw:
        return None
    low = raw.lower()
    if low in HOURS:
        return HOURS[low]
    norm = _norm_header(raw)
    if norm in HOURS:
        return HOURS[norm]
    low_deordinal = low.replace("ª", "a").replace("º", "o")
    if low_deordinal in HOURS:
        return HOURS[low_deordinal]

    try:
        n = int(float(raw.replace(",", ".")))
        if 1 <= n <= 7:
            return n - 1
    except ValueError:
        pass
    return None


def _day_index_from_cell(value: object) -> int | None:
    """
    Día lectivo 0=Lunes … 4=Viernes para la cuadrícula.

    Excel suele devolver:
    - ``datetime`` / ``date`` del propio día (openpyxl): se usa ``weekday()`` (ISO, lunes=0).
    - Enteros o decimales ``1``–``5`` como secuencia Lunes–Viernes.
    - Texto: ``lunes``, ``martes``, … (con o sin tilde), abreviaturas o inglés.
    """
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, datetime):
        wd = value.weekday()
        return wd if wd <= 4 else None

    if isinstance(value, date):
        wd = value.weekday()
        return wd if wd <= 4 else None

    if isinstance(value, (int, float)):
        try:
            n = int(float(value))
        except (TypeError, ValueError):
            return None
        if 1 <= n <= 5:
            return n - 1
        return None

    text = str(value).strip()
    if not text:
        return None

    try:
        n = int(float(text.replace(",", ".")))
        if 1 <= n <= 5:
            return n - 1
    except ValueError:
        pass

    norm = _norm_header(text)
    # Excel ES a menudo guarda abreviaturas con punto final ("lun.", "mar.") que no
    # coinciden con las claves lun/mar; mié–vie suelen ir sin punto o nombre largo.
    tok = norm.strip(".,;:·")
    if tok in DAYS:
        return DAYS[tok]
    if tok in _DAY_ALIASES:
        return _DAY_ALIASES[tok]
    return _DAY_ALIASES.get(norm)


def _teacher_id_from_schedule_label(label: str) -> int | None:
    """
    Usuario activo ya existente en `users` (nombre normalizado, alias, email o parte local).
    No crea cuentas: si no hay coincidencia devuelve None.
    """
    clean = (label or "").strip()
    if not clean:
        return None

    if "@" in clean:
        found = get_user_by_email(clean.lower())
        if found and int(found.get("active") or 0) == 1:
            return int(found["id"])

    teachers = list_teachers_min()
    norm = _norm_header(clean)
    clean_lower = clean.lower()

    for t in teachers:
        tname = _norm_header(t.get("name") or "")
        if tname == norm:
            return int(t["id"])

    for t in teachers:
        alias = (t.get("alias") or "").strip()
        if alias and _norm_header(alias) == norm:
            return int(t["id"])

    for t in teachers:
        email = (t.get("email") or "").strip().lower()
        if email and email == clean_lower:
            return int(t["id"])
        if email:
            local = email.split("@", 1)[0]
            if local == clean_lower:
                return int(t["id"])

    return None


def _note_unknown_teacher_unique(ordered: list[str], seen: set[str], raw_name: str) -> None:
    key = (raw_name or "").strip()
    if key and key not in seen:
        seen.add(key)
        ordered.append(key)


DAY_NAMES_ES = ("lunes", "martes", "miércoles", "jueves", "viernes")
_MAX_CONFLICT_QS = 80


def _hour_label_es(hour_idx: int) -> str:
    if 0 <= hour_idx < len(HOUR_LABELS):
        return HOUR_LABELS[hour_idx]
    return f"hora {hour_idx}"


def _day_label_es(day_idx: int) -> str:
    if 0 <= day_idx < len(DAY_NAMES_ES):
        return DAY_NAMES_ES[day_idx]
    return f"día {day_idx}"


def _incoming_task_phrase(slot_type: str) -> str:
    kind = (slot_type or "").strip().upper()
    if kind == "CLASS":
        return "la clase"
    if kind == "GUARD":
        return "la guardia"
    if kind == "OTHER":
        return "la hora de otros"
    return "la hora"


def _existing_occupancy_phrase(slot: dict) -> str:
    kind = str(slot.get("slot_type") or slot.get("type") or "").strip().upper()
    if kind == "CLASS":
        group = str(slot.get("group") or "").strip()
        subject = str(slot.get("subject") or "").strip()
        if group:
            return f"clase con {group}"
        if subject:
            return f"clase de {subject}"
        return "clase"
    if kind == "GUARD":
        gt = str(slot.get("guard_type") or "").strip()
        return f"guardia {gt}" if gt else "guardia"
    if kind == "OTHER":
        label = str(slot.get("subject") or "").strip()
        return f"otros ({label})" if label else "otros"
    return "otra hora en esa casilla"


def _conflict_import_message(
    *,
    incoming_type: str,
    hour_idx: int,
    day_idx: int,
    teacher_name: str,
    existing: dict,
) -> str:
    who = (teacher_name or "").strip() or "ese profesor"
    return (
        f"No se grabó {_incoming_task_phrase(incoming_type)} de "
        f"{_hour_label_es(hour_idx)} del {_day_label_es(day_idx)} a {who} "
        f"porque ya tiene {_existing_occupancy_phrase(existing)}"
    )


def _recreo_placement_error(
    *,
    slot_type: str,
    hour_index: int,
    day_index: int,
    teacher_name: str,
    guard_type: str | None,
) -> str | None:
    """En recreo solo cabe guardia de recreo; esas guardias no van en otras franjas."""
    who = (teacher_name or "").strip() or "ese profesor"
    kind = (slot_type or "").strip().upper()
    when = f"{_hour_label_es(hour_index)} del {_day_label_es(day_index)} a {who}"
    if hour_index == RECREO_HOUR_INDEX:
        if kind != "GUARD" or not _is_recreo_guard_type(guard_type):
            return (
                f"No se grabó {_incoming_task_phrase(kind)} de {when} "
                f"porque en el recreo solo se admite guardia de recreo"
            )
        return None
    if kind == "GUARD" and _is_recreo_guard_type(guard_type):
        return (
            f"No se grabó la guardia de recreo de {when} "
            f"porque las guardias de recreo solo se admiten en el recreo"
        )
    return None


class _ImportOccupancy:
    """Casillas ya ocupadas (BD + lo grabado en esta misma importación)."""

    def __init__(self) -> None:
        self._by_teacher: dict[int, dict[tuple[int, int], dict]] = {}

    def get(self, teacher_id: int, day_index: int, hour_index: int) -> dict | None:
        return self._map(teacher_id).get((day_index, hour_index))

    def put(self, teacher_id: int, day_index: int, hour_index: int, slot: dict) -> None:
        self._map(teacher_id)[(day_index, hour_index)] = slot

    def _map(self, teacher_id: int) -> dict[tuple[int, int], dict]:
        found = self._by_teacher.get(teacher_id)
        if found is None:
            found = {}
            for s in list_schedule_slots(teacher_id=teacher_id):
                found[(int(s["day_index"]), int(s["hour_index"]))] = dict(s)
            self._by_teacher[teacher_id] = found
        return found


def _try_import_schedule_slot(
    *,
    occupancy: _ImportOccupancy,
    teacher_id: int,
    teacher_name: str,
    day_index: int,
    hour_index: int,
    slot_type: str,
    guard_type: str | None = None,
    group_name: str | None = None,
    room: str | None = None,
    subject: str | None = None,
    source: str,
) -> str | None:
    """Inserta solo si la casilla está libre. Nunca pisa lo ya grabado."""
    placement = _recreo_placement_error(
        slot_type=slot_type,
        hour_index=hour_index,
        day_index=day_index,
        teacher_name=teacher_name,
        guard_type=guard_type,
    )
    if placement:
        return placement
    existing = occupancy.get(teacher_id, day_index, hour_index)
    if existing:
        return _conflict_import_message(
            incoming_type=slot_type,
            hour_idx=hour_index,
            day_idx=day_index,
            teacher_name=teacher_name,
            existing=existing,
        )
    replace_schedule_slot(
        teacher_id=teacher_id,
        day_index=day_index,
        hour_index=hour_index,
        slot_type=slot_type,
        guard_type=guard_type,
        group_name=group_name,
        room=room,
        subject=subject,
        source=source,
    )
    occupancy.put(
        teacher_id,
        day_index,
        hour_index,
        {
            "slot_type": slot_type,
            "guard_type": guard_type,
            "group": group_name,
            "subject": subject,
        },
    )
    return None


def _import_redirect_query(
    *,
    inserted: int,
    skipped: int,
    unknown_teachers: list[str],
    conflicts: list[str],
) -> str:
    q: list[tuple[str, str]] = [
        ("imported", str(inserted)),
        ("skipped", str(skipped)),
    ]
    q.extend(("unknown_teacher", n) for n in unknown_teachers)
    shown = conflicts[:_MAX_CONFLICT_QS]
    q.extend(("conflict", m) for m in shown)
    extra = len(conflicts) - len(shown)
    if extra > 0:
        q.append(("conflict_more", str(extra)))
    return urlencode(q)


def _header_row_indices(ws) -> dict[str, int]:
    headers = [_norm_header(c.value) for c in ws[1]]
    return {h: i for i, h in enumerate(headers) if h}


def _classes_sheet_columns(
    idx: dict[str, int],
) -> tuple[int | None, int | None, int | None, int | None, int | None, int | None]:
    name_i = next((idx[k] for k in ("nombre", "name") if k in idx), None)
    day_i = next((idx[k] for k in ("dia", "día") if k in idx), None)
    hour_i = idx.get("hora")
    if hour_i is None and "franja horaria" in idx:
        hour_i = idx["franja horaria"]
    grupo_i = idx.get("grupo")
    aula_i = idx.get("aula")
    materia_i = idx.get("materia")
    return name_i, day_i, hour_i, grupo_i, aula_i, materia_i


def _classes_headers_ok(idx: dict[str, int]) -> bool:
    cols = _classes_sheet_columns(idx)
    return all(c is not None for c in cols)


def _worksheet_for_classes_import(wb):
    """Primera hoja cuya fila 1 tiene todas las cabeceras; si no, la activa."""
    if _classes_headers_ok(_header_row_indices(wb.active)):
        return wb.active
    for ws in wb.worksheets:
        if ws is wb.active:
            continue
        if _classes_headers_ok(_header_row_indices(ws)):
            return ws
    return wb.active


def _guards_sheet_columns(idx: dict[str, int]) -> tuple[int | None, int | None, int | None, int | None]:
    name_i = next((idx[k] for k in ("nombre", "name") if k in idx), None)
    day_i = next((idx[k] for k in ("dia", "día") if k in idx), None)
    tipo_i = idx.get("tipo")
    if tipo_i is None and "franja horaria" in idx:
        tipo_i = idx["franja horaria"]
    hora_opt_i = idx.get("hora")
    return name_i, day_i, tipo_i, hora_opt_i


def _guards_headers_ok(idx: dict[str, int]) -> bool:
    name_i, day_i, tipo_i, _ = _guards_sheet_columns(idx)
    return name_i is not None and day_i is not None and tipo_i is not None


def _worksheet_for_guards_import(wb):
    if _guards_headers_ok(_header_row_indices(wb.active)):
        return wb.active
    for ws in wb.worksheets:
        if ws is wb.active:
            continue
        if _guards_headers_ok(_header_row_indices(ws)):
            return ws
    return wb.active


def _others_sheet_columns(
    idx: dict[str, int],
) -> tuple[int | None, int | None, int | None, int | None]:
    name_i = next((idx[k] for k in ("nombre", "name") if k in idx), None)
    day_i = next((idx[k] for k in ("dia", "día") if k in idx), None)
    hour_i = idx.get("hora")
    if hour_i is None and "franja horaria" in idx:
        hour_i = idx["franja horaria"]
    known = {"nombre", "name", "dia", "día", "hora", "franja horaria"}
    desc_i = next(
        (
            idx[k]
            for k in ("tarea", "descripcion", "etiqueta", "concepto", "materia", "task")
            if k in idx
        ),
        None,
    )
    if desc_i is None:
        for key, col in idx.items():
            if key in known:
                continue
            if "tarea" in key or key in {"actividad", "task"}:
                desc_i = col
                break
    if desc_i is None:
        used = {i for i in (name_i, day_i, hour_i) if i is not None}
        extras = sorted(i for i in idx.values() if i not in used)
        if extras:
            desc_i = extras[0]
    return name_i, day_i, hour_i, desc_i


def _others_headers_ok(idx: dict[str, int]) -> bool:
    name_i, day_i, hour_i, _ = _others_sheet_columns(idx)
    return name_i is not None and day_i is not None and hour_i is not None


def _worksheet_for_others_import(wb):
    if _others_headers_ok(_header_row_indices(wb.active)):
        return wb.active
    for ws in wb.worksheets:
        if ws is wb.active:
            continue
        if _others_headers_ok(_header_row_indices(ws)):
            return ws
    return wb.active


def _xlsx_only(filename: str | None) -> bool:
    return bool(filename and filename.lower().endswith(".xlsx"))


def _schedule_axis_int(slot: dict, key: str) -> int:
    """Lee day_index / hour_index evitando ``slot.get(k) or -1``: el 0 es válido (lunes, 1ª hora)."""
    v = slot.get(key)
    if v is None:
        return -1
    try:
        return int(v)
    except (TypeError, ValueError):
        return -1


def _slot_for_schedule_template(row: dict) -> dict:
    """
    Misma información que en la app de consulta con SQLAlchemy (`slot.type.name`),
    pero para filas dict/psycopg: la plantilla usa `schedule_kind` (CLASS/GUARD/OTHER).
    """
    out = dict(row)
    raw = out.get("slot_type")
    if raw is None:
        raw = out.get("type")
    if hasattr(raw, "name"):
        raw = raw.name
    out["schedule_kind"] = str(raw or "").strip().upper()
    return out


@router.get("/", response_class=HTMLResponse)
def admin_schedules(
    request: Request,
    teacher_id: int | None = Query(default=None),
    user: dict = Depends(load_user_dep),
):
    _require_permission(user)

    teachers = list_teachers_for_schedule_selector()
    matrix = [[None for _ in range(5)] for _ in range(7)]
    if teacher_id is not None:
        for slot in list_schedule_slots(teacher_id=teacher_id):
            di = _schedule_axis_int(slot, "day_index")
            hi = _schedule_axis_int(slot, "hour_index")
            if 0 <= di <= 4 and 0 <= hi <= 6:
                matrix[hi][di] = _slot_for_schedule_template(slot)

    return _templates(request).TemplateResponse(
        "admin/schedules.html",
        ctx(
            request,
            user=user,
            title="Gestión de horarios",
            teachers=teachers,
            selected_teacher_id=teacher_id,
            schedule=matrix,
            show_schedule=teacher_id is not None,
        ),
        headers={"Cache-Control": "no-store, max-age=0, must-revalidate"},
    )


@router.get("/edit/{teacher_id}", response_class=HTMLResponse)
def admin_schedules_edit_get(
    request: Request,
    teacher_id: int,
    error: str | None = Query(default=None),
    user: dict = Depends(load_user_dep),
):
    _require_permission(user)
    teacher = get_user_by_id(teacher_id)
    if not teacher:
        return RedirectResponse("/admin/schedules/", status_code=303)
    matrix = [[None for _ in range(5)] for _ in range(7)]
    for slot in list_schedule_slots(teacher_id=teacher_id):
        di = _schedule_axis_int(slot, "day_index")
        hi = _schedule_axis_int(slot, "hour_index")
        if 0 <= di <= 4 and 0 <= hi <= 6:
            matrix[hi][di] = _slot_for_schedule_template(slot)

    groups = list_groups()
    extra_groups: set[str] = set()
    for row in matrix:
        for slot in row:
            if not slot:
                continue
            g = str(slot.get("group") or "").strip()
            if g and g.casefold() not in {x.casefold() for x in groups}:
                extra_groups.add(g)
    if extra_groups:
        groups = sorted({*groups, *extra_groups}, key=normalize_for_sort)
    materias_por_grupo = map_materias_horario_por_grupo()

    return _templates(request).TemplateResponse(
        "admin/schedules_edit.html",
        ctx(
            request,
            user=user,
            title=f"Editar horario · {teacher.get('name') or ''}",
            teacher=teacher,
            schedule=matrix,
            groups=groups,
            materias_por_grupo_json=json.dumps(
                materias_por_grupo, ensure_ascii=False
            ).replace("<", "\\u003c"),
            guard_labels=sorted(GUARD_LABELS, key=normalize_for_sort),
            form_error=(error or "").strip(),
        ),
        headers={"Cache-Control": "no-store, max-age=0, must-revalidate"},
    )


@router.post("/edit/{teacher_id}")
async def admin_schedules_edit_post(
    request: Request,
    teacher_id: int,
    user: dict = Depends(load_user_dep),
):
    _require_permission(user)
    teacher = get_user_by_id(teacher_id)
    if not teacher:
        return RedirectResponse("/admin/schedules/", status_code=303)

    flat = await _read_schedule_post_flat(request)
    if not flat or not _schedule_form_has_grid(flat):
        return RedirectResponse(
            f"/admin/schedules/edit/{teacher_id}?error=incomplete_form",
            status_code=303,
        )

    cells = _build_schedule_cells_from_flat(flat)
    expected_slots = sum(
        1
        for c in cells
        if (
            c.get("kind") == "CLASS"
            and (
                (str(c.get("group") or "").strip())
                or (str(c.get("subject") or "").strip())
            )
        )
        or (
            c.get("kind") == "GUARD"
            and (str(c.get("guard_type") or "").strip())
        )
        or c.get("kind") == "OTHER"
    )
    apply_teacher_schedule_grid_edits(teacher_id=teacher_id, cells=cells)
    rows_after = list_schedule_slots(teacher_id=teacher_id)
    slot_n = len(rows_after)
    if expected_slots > 0 and slot_n == 0:
        _log.error(
            "Horario teacher_id=%s: %s celdas válidas en POST pero 0 filas tras guardar",
            teacher_id,
            expected_slots,
        )
        return RedirectResponse(
            f"/admin/schedules/edit/{teacher_id}?error=save_failed",
            status_code=303,
        )
    class_n = sum(
        1 for r in rows_after if str(r.get("slot_type") or "").strip().upper() == "CLASS"
    )
    guard_n = sum(
        1 for r in rows_after if str(r.get("slot_type") or "").strip().upper() == "GUARD"
    )
    other_n = sum(
        1 for r in rows_after if str(r.get("slot_type") or "").strip().upper() == "OTHER"
    )

    try:
        add_action_log(
            user_id=user.get("id"),
            action="schedule_edit_save",
            entity="schedule_slots",
            entity_id=teacher_id,
            detail=(
                f"Horario editado manualmente teacher_id={teacher_id} slots={slot_n} "
                f"class={class_n} guard={guard_n} other={other_n}"
            ),
        )
    except Exception:
        _log.exception("add_action_log falló tras guardar horario; los cambios en schedule_slots ya están confirmados")

    q = urlencode(
        {
            "teacher_id": str(teacher_id),
            "saved": "1",
            "slots": str(slot_n),
            "class_slots": str(class_n),
            "guard_slots": str(guard_n),
            "other_slots": str(other_n),
        }
    )
    dest = f"/admin/schedules/?{q}"
    return RedirectResponse(
        dest,
        status_code=303,
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate",
            "Access-Control-Expose-Headers": "Location",
        },
    )


@router.get("/export.pdf")
def admin_schedules_export_pdf(
    teacher_id: int = Query(...),
    user: dict = Depends(load_user_dep),
):
    _require_permission(user)

    teachers = {int(t["id"]): str(t["name"]) for t in list_teachers_min()}
    if teacher_id not in teachers:
        return RedirectResponse("/admin/schedules/", status_code=303)
    teacher_name = teachers[teacher_id]

    matrix = [[None for _ in range(5)] for _ in range(7)]
    for slot in list_schedule_slots(teacher_id=teacher_id):
        di = _schedule_axis_int(slot, "day_index")
        hi = _schedule_axis_int(slot, "hour_index")
        if 0 <= di <= 4 and 0 <= hi <= 6:
            st = (slot.get("slot_type") or "").strip().upper()
            if st == "CLASS":
                matrix[hi][di] = {
                    "type": "CLASS",
                    "group": slot.get("group") or "",
                    "room": slot.get("room") or "",
                    "subject": slot.get("subject") or "",
                }
            elif st == "GUARD":
                matrix[hi][di] = {
                    "type": "GUARD",
                    "guard_type": slot.get("guard_type") or "",
                }
            elif st == "OTHER":
                matrix[hi][di] = {
                    "type": "OTHER",
                    "subject": slot.get("subject") or "Otros",
                }

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    tmp.close()
    generate_schedule_pdf(
        path=tmp.name,
        teacher_name=teacher_name,
        center_name="IES",
        schedule=matrix,
    )
    fn = safe_pdf_filename(f"horario_{teacher_name or teacher_id}", ext="pdf")
    return pdf_attachment_response(tmp.name, filename=fn)


@router.get("/imports", response_class=HTMLResponse)
def admin_schedules_imports(request: Request, user: dict = Depends(load_user_dep)):
    _require_permission(user)
    return _templates(request).TemplateResponse(
        "admin/schedules_imports.html",
        ctx(request, user=user, title="Gestión de horarios · Importaciones"),
    )


@router.get("/imports/classes", response_class=HTMLResponse)
def admin_schedules_import_classes_get(request: Request, user: dict = Depends(load_user_dep)):
    _require_permission(user)
    return _templates(request).TemplateResponse(
        "admin/schedules_import_classes.html",
        ctx(request, user=user, title="Importar horas de clase"),
    )


@router.post("/imports/classes")
def admin_schedules_import_classes_post(
    file: UploadFile = File(...),
    user: dict = Depends(load_user_dep),
):
    _require_permission(user)
    if not _xlsx_only(file.filename):
        return RedirectResponse("/admin/schedules/imports/classes?error=format", status_code=303)
    try:
        wb = openpyxl.load_workbook(file.file)
        ws = _worksheet_for_classes_import(wb)
    except Exception:
        return RedirectResponse("/admin/schedules/imports/classes?error=parse", status_code=303)

    idx = _header_row_indices(ws)
    name_i, day_i, hour_i, grupo_i, aula_i, materia_i = _classes_sheet_columns(idx)
    if name_i is None or day_i is None or hour_i is None:
        return RedirectResponse("/admin/schedules/imports/classes?error=columns", status_code=303)
    if grupo_i is None or aula_i is None or materia_i is None:
        return RedirectResponse("/admin/schedules/imports/classes?error=columns", status_code=303)

    inserted = 0
    skipped = 0
    unknown_teachers: list[str] = []
    unknown_seen: set[str] = set()
    conflicts: list[str] = []
    occupancy = _ImportOccupancy()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = str(row[name_i]).strip() if name_i < len(row) and row[name_i] is not None else ""
        day_cell = row[day_i] if day_i < len(row) else None
        day_idx = _day_index_from_cell(day_cell)
        hour_cell = row[hour_i] if hour_i < len(row) else None
        if not name or day_idx is None:
            skipped += 1
            continue
        hour_idx = _hour_cell_index(hour_cell)
        if hour_idx is None:
            skipped += 1
            continue
        group = str(row[grupo_i]).strip() if grupo_i < len(row) and row[grupo_i] is not None else ""
        room = str(row[aula_i]).strip() if aula_i < len(row) and row[aula_i] is not None else ""
        subject = str(row[materia_i]).strip() if materia_i < len(row) and row[materia_i] is not None else ""
        tid = _teacher_id_from_schedule_label(name)
        if tid is None:
            skipped += 1
            _note_unknown_teacher_unique(unknown_teachers, unknown_seen, name)
            continue
        conflict = _try_import_schedule_slot(
            occupancy=occupancy,
            teacher_id=tid,
            teacher_name=name,
            day_index=day_idx,
            hour_index=hour_idx,
            slot_type="CLASS",
            group_name=group or None,
            room=room or None,
            subject=subject or None,
            source="classes_excel",
        )
        if conflict:
            conflicts.append(conflict)
            continue
        inserted += 1

    add_action_log(
        user_id=user.get("id"),
        action="import_classes",
        entity="schedule_slots",
        detail=(
            f"Import classes: inserted={inserted}, skipped={skipped}, "
            f"conflicts={len(conflicts)}, unknown_teachers={len(unknown_teachers)}"
        ),
    )
    return RedirectResponse(
        "/admin/schedules/imports/classes?"
        + _import_redirect_query(
            inserted=inserted,
            skipped=skipped,
            unknown_teachers=unknown_teachers,
            conflicts=conflicts,
        ),
        status_code=303,
    )


@router.get("/imports/guards", response_class=HTMLResponse)
def admin_schedules_import_guards_get(request: Request, user: dict = Depends(load_user_dep)):
    _require_permission(user)
    return _templates(request).TemplateResponse(
        "admin/schedules_import_guards.html",
        ctx(request, user=user, title="Importar guardias"),
    )


@router.post("/imports/guards")
def admin_schedules_import_guards_post(
    file: UploadFile = File(...),
    user: dict = Depends(load_user_dep),
):
    _require_permission(user)
    if not _xlsx_only(file.filename):
        return RedirectResponse("/admin/schedules/imports/guards?error=format", status_code=303)
    try:
        wb = openpyxl.load_workbook(file.file)
        ws = _worksheet_for_guards_import(wb)
    except Exception:
        return RedirectResponse("/admin/schedules/imports/guards?error=parse", status_code=303)

    idx = _header_row_indices(ws)
    name_i, day_i, tipo_i, hora_opt_i = _guards_sheet_columns(idx)
    if name_i is None or day_i is None or tipo_i is None:
        return RedirectResponse("/admin/schedules/imports/guards?error=columns", status_code=303)

    inserted = 0
    skipped = 0
    unknown_teachers: list[str] = []
    unknown_seen: set[str] = set()
    conflicts: list[str] = []
    occupancy = _ImportOccupancy()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = str(row[name_i]).strip() if name_i < len(row) and row[name_i] is not None else ""
        day_cell = row[day_i] if day_i < len(row) else None
        day_idx = _day_index_from_cell(day_cell)
        tipo = str(row[tipo_i]).strip() if tipo_i < len(row) and row[tipo_i] is not None else ""
        if not name or day_idx is None or not tipo:
            skipped += 1
            continue
        if tipo not in GUARD_LABELS:
            skipped += 1
            continue
        hour_idx: int | None = RECREO_HOUR_INDEX if tipo.startswith("G RECREO") else None
        if not tipo.startswith("G RECREO") and hora_opt_i is not None and hora_opt_i < len(row) and row[hora_opt_i] is not None:
            parsed_h = _hour_cell_index(row[hora_opt_i])
            if parsed_h is not None:
                hour_idx = parsed_h
        if day_idx is None or hour_idx is None:
            skipped += 1
            continue
        tid = _teacher_id_from_schedule_label(name)
        if tid is None:
            skipped += 1
            _note_unknown_teacher_unique(unknown_teachers, unknown_seen, name)
            continue
        conflict = _try_import_schedule_slot(
            occupancy=occupancy,
            teacher_id=tid,
            teacher_name=name,
            day_index=day_idx,
            hour_index=hour_idx,
            slot_type="GUARD",
            guard_type=tipo,
            source="guards_excel",
        )
        if conflict:
            conflicts.append(conflict)
            continue
        inserted += 1

    add_action_log(
        user_id=user.get("id"),
        action="import_guards",
        entity="schedule_slots",
        detail=(
            f"Import guards: inserted={inserted}, skipped={skipped}, "
            f"conflicts={len(conflicts)}, unknown_teachers={len(unknown_teachers)}"
        ),
    )
    return RedirectResponse(
        "/admin/schedules/imports/guards?"
        + _import_redirect_query(
            inserted=inserted,
            skipped=skipped,
            unknown_teachers=unknown_teachers,
            conflicts=conflicts,
        ),
        status_code=303,
    )


@router.get("/imports/others", response_class=HTMLResponse)
def admin_schedules_import_others_get(request: Request, user: dict = Depends(load_user_dep)):
    _require_permission(user)
    return _templates(request).TemplateResponse(
        "admin/schedules_import_others.html",
        ctx(request, user=user, title="Importar otras horas"),
    )


@router.post("/imports/others")
def admin_schedules_import_others_post(
    file: UploadFile = File(...),
    user: dict = Depends(load_user_dep),
):
    _require_permission(user)
    if not _xlsx_only(file.filename):
        return RedirectResponse("/admin/schedules/imports/others?error=format", status_code=303)
    try:
        wb = openpyxl.load_workbook(file.file)
        ws = _worksheet_for_others_import(wb)
    except Exception:
        return RedirectResponse("/admin/schedules/imports/others?error=parse", status_code=303)

    idx = _header_row_indices(ws)
    name_i, day_i, hour_i, desc_i = _others_sheet_columns(idx)
    if name_i is None or day_i is None or hour_i is None:
        return RedirectResponse("/admin/schedules/imports/others?error=columns", status_code=303)

    inserted = 0
    skipped = 0
    unknown_teachers: list[str] = []
    unknown_seen: set[str] = set()
    conflicts: list[str] = []
    occupancy = _ImportOccupancy()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = str(row[name_i]).strip() if name_i < len(row) and row[name_i] is not None else ""
        day_cell = row[day_i] if day_i < len(row) else None
        day_idx = _day_index_from_cell(day_cell)
        hour_cell = row[hour_i] if hour_i < len(row) else None
        hour_idx = _hour_cell_index(hour_cell)
        if not name or day_idx is None or hour_idx is None:
            skipped += 1
            continue
        label = ""
        if desc_i is not None and desc_i < len(row) and row[desc_i] is not None:
            label = str(row[desc_i]).strip()
        tid = _teacher_id_from_schedule_label(name)
        if tid is None:
            skipped += 1
            _note_unknown_teacher_unique(unknown_teachers, unknown_seen, name)
            continue
        conflict = _try_import_schedule_slot(
            occupancy=occupancy,
            teacher_id=tid,
            teacher_name=name,
            day_index=day_idx,
            hour_index=hour_idx,
            slot_type="OTHER",
            subject=label or None,
            source="others_excel",
        )
        if conflict:
            conflicts.append(conflict)
            continue
        inserted += 1

    add_action_log(
        user_id=user.get("id"),
        action="import_others",
        entity="schedule_slots",
        detail=(
            f"Import others: inserted={inserted}, skipped={skipped}, "
            f"conflicts={len(conflicts)}, unknown_teachers={len(unknown_teachers)}"
        ),
    )
    return RedirectResponse(
        "/admin/schedules/imports/others?"
        + _import_redirect_query(
            inserted=inserted,
            skipped=skipped,
            unknown_teachers=unknown_teachers,
            conflicts=conflicts,
        ),
        status_code=303,
    )


@router.post("/imports/teachers")
def admin_schedules_import_teachers(
    file: UploadFile = File(...),
    user: dict = Depends(load_user_dep),
):
    _require_permission(user)
    if not _xlsx_only(file.filename):
        return RedirectResponse("/admin/schedules/imports?status=error_teachers", status_code=303)
    try:
        wb = openpyxl.load_workbook(file.file)
        ws = wb.active
    except Exception:
        return RedirectResponse("/admin/schedules/imports?status=error_teachers", status_code=303)

    headers = [_norm_header(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h}
    name_i = next((idx[k] for k in ("nombre", "name") if k in idx), None)
    email_i = next((idx[k] for k in ("email", "correo") if k in idx), None)
    if name_i is None or email_i is None:
        return RedirectResponse("/admin/schedules/imports?status=error_teachers", status_code=303)

    created = 0
    updated = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = row[name_i] if name_i < len(row) else None
        email = row[email_i] if email_i < len(row) else None
        if not name or not email:
            skipped += 1
            continue
        alias_v = row[idx["alias"]] if "alias" in idx and idx["alias"] < len(row) else None
        role_v = row[idx["rol"]] if "rol" in idx and idx["rol"] < len(row) else None
        status_v = row[idx["status"]] if "status" in idx and idx["status"] < len(row) else None
        try:
            res = upsert_teacher_from_import(
                name=str(name),
                email=str(email),
                alias=str(alias_v) if alias_v is not None else None,
                role=str(role_v) if role_v is not None else None,
                status=str(status_v) if status_v is not None else None,
            )
            if res == "created":
                created += 1
            else:
                updated += 1
        except ValueError:
            skipped += 1
    add_action_log(
        user_id=user.get("id"),
        action="import_teachers",
        entity="users",
        detail=f"Import teachers: created={created}, updated={updated}, skipped={skipped}",
    )
    return RedirectResponse(
        f"/admin/schedules/imports?status=ok_teachers&created={created}&updated={updated}&skipped={skipped}",
        status_code=303,
    )
