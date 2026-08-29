# routers/admin_students.py

import io
import logging
import unicodedata
from urllib.parse import quote_plus

from utils.local_deps import ensure_local_deps

ensure_local_deps()
try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response

from auth import load_user_dep
from context import ctx
from db.groups import list_groups
from db.students import (
    create_student_manual,
    ensure_students_schema,
    get_all_groups,
    get_all_students,
    get_student_by_id,
    parse_bool_import,
    parse_date_import,
    update_student_admin,
    upsert_student_from_import,
)
from utils.enums import PERM_GESTION_ALUMNOS
from utils.permissions import has_permission
from utils.text import normalize_for_sort

router = APIRouter()
_log = logging.getLogger(__name__)
_DEFAULT_PAGE_SIZE = 20
_PAGE_SIZE_OPTIONS = (20, 50, 100)


# ------------------------------------------------------
# Permiso
# ------------------------------------------------------

def _require_perm(user: dict):
    if not has_permission(user, PERM_GESTION_ALUMNOS):
        raise HTTPException(status_code=403)


def _resolve_per_page(raw: str | None) -> tuple[int | None, str]:
    """Devuelve (tamaño o None=todos, clave UI)."""
    text = (raw or "").strip().lower()
    if text in {"all", "todos", "todo"}:
        return None, "all"
    if text.isdigit() and int(text) in _PAGE_SIZE_OPTIONS:
        return int(text), text
    return _DEFAULT_PAGE_SIZE, str(_DEFAULT_PAGE_SIZE)


def _students_list_url(
    *,
    status: str | None = None,
    msg: str | None = None,
    grupo: str | None = None,
    page: int | None = None,
    per_page: str | None = None,
    edit: int | None = None,
) -> str:
    parts: list[str] = []
    if status:
        parts.append(f"status={status}")
    if msg:
        parts.append(f"msg={quote_plus(msg)}")
    if grupo:
        parts.append(f"grupo={quote_plus(grupo)}")
    if per_page and per_page != str(_DEFAULT_PAGE_SIZE):
        parts.append(f"per_page={quote_plus(per_page)}")
    if page and page > 1:
        parts.append(f"page={page}")
    if edit is not None:
        parts.append(f"edit={edit}")
    qs = "&".join(parts)
    return f"/admin/students?{qs}" if qs else "/admin/students"


def _parse_optional_bool_form(raw: str | None) -> bool | None:
    text = (raw or "").strip()
    if not text:
        return None
    return parse_bool_import(text)


def _import_redirect(status: str, *, msg: str | None = None) -> RedirectResponse:
    qs = f"status={status}"
    if msg:
        qs += f"&msg={quote_plus(msg)}"
    return RedirectResponse(f"/admin/students?{qs}", status_code=303)


def _norm_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = "".join(
        c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn"
    )
    return " ".join(text.split())


def _load_uploaded_workbook(file: UploadFile):
    """Lee el Excel subido en memoria (más fiable que pasar el stream directo)."""
    raw = file.file.read()
    if not raw:
        raise ValueError("Archivo vacío")
    if raw[:2] != b"PK":
        raise ValueError("No es un .xlsx válido (formato ZIP/OpenXML)")
    return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)


def _has_student_name_headers(labels: list[str]) -> bool:
    if any(x in labels for x in ("alumno", "datos")):
        return True
    return "nombre" in labels and ("apellido1" in labels or "apellido2" in labels)


def _header_row_index(ws) -> int:
    for row_num in range(1, min(ws.max_row or 1, 6) + 1):
        labels = [_norm_header(cell.value) for cell in ws[row_num]]
        if "grupo" in labels and _has_student_name_headers(labels):
            return row_num
    return 1


def _build_import_column_index(headers: list[object]) -> dict[str, int] | None:
    idx: dict[str, int] = {}
    for pos, raw in enumerate(headers):
        norm = _norm_header(raw)
        if not norm:
            continue
        if norm == "grupo" and "Grupo" not in idx:
            idx["Grupo"] = pos
        elif norm in {"alumno", "datos", "nombre alumno", "alumnos", "nombre del alumno"} and "Alumno" not in idx:
            idx["Alumno"] = pos
        elif norm == "apellido1":
            idx["Apellido1"] = pos
        elif norm == "apellido2":
            idx["Apellido2"] = pos
        elif norm == "nombre" and "Nombre" not in idx:
            idx["Nombre"] = pos
        elif norm in {"sexo", "genero"}:
            idx["Sexo"] = pos
        elif norm in {"email alumno", "email del alumno", "correo alumno", "e_mail", "email"}:
            idx["Email alumno"] = pos
        elif norm in {"email madre", "correo madre"}:
            idx["Email madre"] = pos
        elif norm in {"email padre", "correo padre"}:
            idx["Email padre"] = pos
        elif norm == "cie":
            idx["CIE"] = pos
        elif norm in {"doc", "dni", "documento", "doc. alumno", "doc alumno"}:
            idx["Doc"] = pos
        elif norm in {
            "fecha nacimiento",
            "fecha de nacimiento",
            "f. nacimiento",
            "f_nacimiento",
        }:
            idx["Fecha nacimiento"] = pos
        elif norm in {"telefono 1", "telefono1", "tfno 1", "tel 1"}:
            idx["Teléfono 1"] = pos
        elif norm in {"telefono 2", "telefono2", "tfno 2", "tel 2"}:
            idx["Teléfono 2"] = pos
        elif norm in {"obs. telefonos", "obs telefonos", "obs. teléfonos", "observaciones telefonos"}:
            idx["Obs. teléfonos"] = pos
        elif norm in {"difusion imagen", "difusion de imagen"}:
            idx["Difusión imagen"] = pos
        elif norm == "transporte":
            idx["Transporte"] = pos
        elif norm in {"repite curso", "repetidor", "repite"} or (
            "repite" in norm and "curso" in norm
        ):
            idx["Repetidor"] = pos
        elif norm in {"parada", "t_parada"}:
            idx["Parada"] = pos
    if "Grupo" not in idx:
        return None
    if "Alumno" not in idx and not (
        "Nombre" in idx and ("Apellido1" in idx or "Apellido2" in idx)
    ):
        return None
    return idx


def _cell(row: tuple, index: int):
    if index >= len(row):
        return None
    return row[index]


def _student_name_from_row(row: tuple, idx: dict[str, int]) -> str:
    if "Alumno" in idx:
        val = _cell(row, idx["Alumno"])
        text = str(val).strip() if val is not None else ""
        if text:
            return text

    apellidos: list[str] = []
    for key in ("Apellido1", "Apellido2"):
        if key in idx:
            val = _cell(row, idx[key])
            if val is not None and str(val).strip():
                apellidos.append(str(val).strip())

    nombre = ""
    if "Nombre" in idx:
        val = _cell(row, idx["Nombre"])
        if val is not None and str(val).strip():
            nombre = str(val).strip()

    if apellidos and nombre:
        return f"{' '.join(apellidos)}, {nombre}"
    if apellidos:
        return " ".join(apellidos)
    return nombre


def _normalize_import_sexo(value: object) -> str | None:
    sx = _norm_header(value)
    if not sx:
        return None
    mapping = {
        "m": "M",
        "v": "V",
        "f": "M",
        "mujer": "M",
        "muj": "M",
        "femenino": "M",
        "h": "V",
        "hombre": "V",
        "varon": "V",
        "masculino": "V",
    }
    return mapping.get(sx, sx.upper() if sx.upper() in ("M", "V") else None)


def _parse_repite_curso_value(value: object) -> bool | None:
    """Valores de la columna REPITE CURSO del Excel del centro."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if value == 1:
            return True
        if value == 0:
            return False
        return None
    text = _norm_header(value)
    if not text:
        return None
    if text in {"si", "s", "1", "true", "yes", "y", "x", "rep", "repite", "repetidor"}:
        return True
    if text in {"no", "n", "0", "false", "falso"}:
        return False
    try:
        return parse_bool_import(value)
    except ValueError:
        return None


def _set_bool_import_field(
    kwargs: dict,
    row: tuple,
    idx: dict[str, int],
    col_key: str,
    field_name: str,
    *,
    parser=parse_bool_import,
) -> None:
    if col_key not in idx:
        return
    v = _cell(row, idx[col_key])
    if v is None:
        return
    if isinstance(v, str) and not v.strip():
        return
    parsed = parser(v)
    if parsed is not None:
        kwargs[field_name] = parsed


# ------------------------------------------------------
# Vista principal
# ------------------------------------------------------

@router.get("/admin/students", response_class=HTMLResponse)
def admin_students(
    request: Request,
    user: dict = Depends(load_user_dep),
    grupo: str | None = None,
    page: int = 1,
):
    _require_perm(user)
    ensure_students_schema()

    groups = get_all_groups()
    selected_group = (grupo or "").strip() or None
    # Leer per_page de la query (por defecto 20)
    page_size, per_page_key = _resolve_per_page(request.query_params.get("per_page"))

    if selected_group:
        students_all = [
            s for s in get_all_students()
            if s["grupo"] == selected_group
        ]
    else:
        students_all = get_all_students()

    total = len(students_all)
    if page_size is None:
        # Todos
        total_pages = 1
        page = 1
        start = 0
        end = total
        students = students_all
        effective_size = total if total else 1
    else:
        total_pages = max(1, (total + page_size - 1) // page_size) if total else 1
        try:
            page_num = int(page or 1)
        except (TypeError, ValueError):
            page_num = 1
        page = max(1, min(page_num, total_pages))
        start = (page - 1) * page_size
        end = start + page_size
        students = students_all[start:end]
        effective_size = page_size

    showing_from = (start + 1) if total else 0
    showing_to = min(end, total) if total else 0

    edit_student = None
    raw_edit = (request.query_params.get("edit") or "").strip()
    if raw_edit.isdigit():
        edit_student = get_student_by_id(int(raw_edit))

    catalog_groups = list_groups()

    return request.app.state.templates.TemplateResponse(
        "admin/students.html",
        ctx(
            request,
            user=user,
            title="Gestión de alumnos",
            students=students,
            groups=groups,
            catalog_groups=catalog_groups,
            selected_group=selected_group,
            edit_student=edit_student,
            page=page,
            total_pages=total_pages,
            total_students=total,
            page_size=effective_size,
            per_page=per_page_key,
            showing_from=showing_from,
            showing_to=showing_to,
            show_all=(per_page_key == "all"),
        ),
    )


@router.post("/admin/students/create")
def admin_students_create(
    user: dict = Depends(load_user_dep),
    grupo: str = Form(...),
    alumno: str = Form(...),
    sexo: str = Form(...),
):
    _require_perm(user)

    grupo_q = quote_plus((grupo or "").strip())
    try:
        create_student_manual(grupo=grupo, alumno=alumno, sexo=sexo)
    except ValueError as exc:
        msg = quote_plus(str(exc))
        return RedirectResponse(
            f"/admin/students?status=create_error&msg={msg}&grupo={grupo_q}",
            status_code=303,
        )

    return RedirectResponse(
        f"/admin/students?status=created_one&grupo={grupo_q}",
        status_code=303,
    )


@router.post("/admin/students/update/{student_id}")
def admin_students_update(
    student_id: int,
    user: dict = Depends(load_user_dep),
    grupo: str = Form(...),
    alumno: str = Form(...),
    sexo: str = Form(""),
    email_student: str = Form(""),
    email_mother: str = Form(""),
    email_father: str = Form(""),
    cie: str = Form(""),
    doc: str = Form(""),
    fecha_nacimiento: str = Form(""),
    telefono1: str = Form(""),
    telefono2: str = Form(""),
    obs_tfno: str = Form(""),
    difusion_imagen: str = Form(""),
    transporte: str = Form(""),
    repetidor: str = Form(""),
    parada: str = Form(""),
    return_grupo: str = Form(""),
    return_page: int = Form(1),
    return_per_page: str = Form("20"),
):
    _require_perm(user)

    grupo_clean = (grupo or "").strip()
    ret_grupo = (return_grupo or "").strip() or None
    ret_page = max(1, int(return_page or 1))
    _, ret_per_page = _resolve_per_page(return_per_page)

    try:
        fecha = parse_date_import(fecha_nacimiento) if (fecha_nacimiento or "").strip() else None
        update_student_admin(
            student_id=student_id,
            grupo=grupo_clean,
            alumno=alumno,
            sexo=sexo or None,
            email_student=email_student,
            email_mother=email_mother,
            email_father=email_father,
            cie=cie,
            doc=doc,
            fecha_nacimiento=fecha,
            telefono1=telefono1,
            telefono2=telefono2,
            obs_tfno=obs_tfno,
            difusion_imagen=_parse_optional_bool_form(difusion_imagen),
            transporte=_parse_optional_bool_form(transporte),
            repetidor=_parse_optional_bool_form(repetidor),
            parada=parada,
        )
    except ValueError as exc:
        return RedirectResponse(
            _students_list_url(
                status="update_error",
                msg=str(exc),
                grupo=ret_grupo,
                page=ret_page,
                per_page=ret_per_page,
                edit=student_id,
            ),
            status_code=303,
        )

    return RedirectResponse(
        _students_list_url(
            status="updated",
            grupo=ret_grupo,
            page=ret_page,
            per_page=ret_per_page,
        ),
        status_code=303,
    )


# ------------------------------------------------------
# Exportar Excel
# ------------------------------------------------------

def _student_headers() -> list[str]:
    return [
        "Grupo",
        "Alumno",
        "Sexo",
        "Email alumno",
        "Email madre",
        "Email padre",
        "CIE",
        "Doc",
        "Fecha nacimiento",
        "Teléfono 1",
        "Teléfono 2",
        "Obs. teléfonos",
        "Difusión imagen",
        "Transporte",
        "Repetidor",
        "Parada",
    ]


def _student_export_row(s: dict) -> list:
    return [
        s["grupo"],
        s["alumno"],
        s.get("sexo") or "",
        s.get("email_student") or "",
        s.get("email_mother") or "",
        s.get("email_father") or "",
        s.get("cie") or "",
        s.get("doc") or "",
        s.get("fecha_nacimiento_display") or "",
        s.get("telefono1") or "",
        s.get("telefono2") or "",
        s.get("obs_tfno") or "",
        s.get("difusion_imagen_display") or "",
        s.get("transporte_display") or "",
        s.get("repetidor_display") or "",
        s.get("parada") or "",
    ]


def _apply_optional_import_columns(
    kwargs: dict, idx: dict[str, int], row: tuple
) -> None:
    if "Sexo" in idx and idx["Sexo"] < len(row):
        v = row[idx["Sexo"]]
        if v is not None and str(v).strip():
            sx = _normalize_import_sexo(v)
            if sx:
                kwargs["sexo"] = sx

    for col_key, arg_name in (
        ("Email alumno", "email_student"),
        ("Email madre", "email_mother"),
        ("Email padre", "email_father"),
        ("CIE", "cie"),
        ("Doc", "doc"),
        ("Teléfono 1", "telefono1"),
        ("Teléfono 2", "telefono2"),
        ("Obs. teléfonos", "obs_tfno"),
        ("Parada", "parada"),
    ):
        if col_key in idx:
            v = _cell(row, idx[col_key])
            if v is not None and str(v).strip():
                kwargs[arg_name] = str(v).strip()

    if "Fecha nacimiento" in idx:
        v = _cell(row, idx["Fecha nacimiento"])
        if v is not None and str(v).strip():
            kwargs["fecha_nacimiento"] = parse_date_import(v)

    _set_bool_import_field(kwargs, row, idx, "Difusión imagen", "difusion_imagen")
    _set_bool_import_field(kwargs, row, idx, "Transporte", "transporte")
    _set_bool_import_field(
        kwargs, row, idx, "Repetidor", "repetidor", parser=_parse_repite_curso_value
    )


@router.get("/admin/students/export")
def export_students(
    user: dict = Depends(load_user_dep),
):
    _require_perm(user)

    students = get_all_students()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumnos"

    ws.append(_student_headers())

    for s in students:
        ws.append(_student_export_row(s))

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return Response(
        stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=alumnos.xlsx"
        },
    )


# ------------------------------------------------------
# Importar Excel
# ------------------------------------------------------

@router.post("/admin/students/import")
def import_students(
    user: dict = Depends(load_user_dep),
    file: UploadFile = File(...),
):
    _require_perm(user)
    ensure_students_schema()

    filename = (file.filename or "").strip().lower()
    if filename and not filename.endswith((".xlsx", ".xlsm")):
        return _import_redirect("error", msg="Use un archivo Excel .xlsx")

    try:
        wb = _load_uploaded_workbook(file)
        ws = wb.active
    except ValueError as exc:
        _log.warning("Import alumnos rechazado: %s", exc)
        return _import_redirect("error", msg=str(exc))
    except Exception:
        _log.exception("Import alumnos: fallo al leer Excel")
        return _import_redirect(
            "error",
            msg="No se pudo leer el Excel. Compruebe que es .xlsx (no .xls) y no está dañado.",
        )

    header_row = _header_row_index(ws)
    headers = [cell.value for cell in ws[header_row]]
    idx = _build_import_column_index(headers)
    if not idx:
        found = ", ".join(str(h).strip() for h in headers if h is not None and str(h).strip()) or "—"
        return _import_redirect(
            "error",
            msg=(
                "Faltan Grupo y nombre de alumno (columna Alumno, DATOS o "
                f"APELLIDO1/APELLIDO2+NOMBRE). Cabeceras detectadas: {found}"
            ),
        )

    created = 0
    updated = 0
    skipped = 0
    invalid_groups = 0
    invalid_group_names: set[str] = set()
    valid_groups_norm = {g.strip().lower(): g for g in list_groups()}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        grupo_raw = _cell(row, idx["Grupo"])
        alumno = _student_name_from_row(row, idx)
        if grupo_raw is None or not alumno:
            continue
        grupo = str(grupo_raw).strip()
        if not grupo:
            continue
        if grupo.lower() not in valid_groups_norm:
            invalid_groups += 1
            skipped += 1
            invalid_group_names.add(grupo)
            continue

        kwargs: dict = {"grupo": valid_groups_norm[grupo.lower()], "alumno": alumno}

        try:
            _apply_optional_import_columns(kwargs, idx, row)
        except ValueError:
            skipped += 1
            continue

        try:
            res = upsert_student_from_import(**kwargs)
        except ValueError:
            skipped += 1
            continue

        if res == "created":
            created += 1
        elif res == "updated":
            updated += 1

    if created == 0 and updated == 0 and skipped == 0:
        return _import_redirect(
            "error",
            msg="No se encontraron filas válidas con Grupo y Alumno informados.",
        )

    status = "imported_invalid" if invalid_groups > 0 else "imported"
    invalid_groups_list = ",".join(sorted(invalid_group_names, key=normalize_for_sort))
    invalid_groups_list_q = quote_plus(invalid_groups_list)
    return RedirectResponse(
        f"/admin/students?status={status}&created={created}&updated={updated}&skipped={skipped}&invalid_groups={invalid_groups}&invalid_group_names={invalid_groups_list_q}",
        status_code=303,
    )
