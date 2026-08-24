"""Administración: catálogo curso de asignatura (importación Excel)."""

from __future__ import annotations

import io
import logging
from urllib.parse import quote

from utils.local_deps import ensure_local_deps

ensure_local_deps()
try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]
from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response

from auth import load_user_dep
from context import ctx
from db.enrolled_subject_catalog import (
    CATALOG_COURSE_NUMS,
    CATALOG_ETAPA_EXPORT_LABELS,
    CATALOG_EXCEL_HEADERS,
    CATALOG_FIELD_NAMES,
    CATALOG_STAGE_KEYS,
    get_catalog_meta,
    list_catalog_for_export,
    list_catalog_preview,
    normalize_catalog_etapa,
    parse_catalog_workbook_rows,
    replace_subject_catalog,
)
from utils.enums import PERM_ASIGNATURAS_MATRICULADAS
from utils.permissions import has_permission

router = APIRouter(
    prefix="/admin/catalogo-asignaturas",
    tags=["admin_subject_catalog"],
)

_log = logging.getLogger(__name__)


def _require_perm(user: dict) -> None:
    if not has_permission(user, PERM_ASIGNATURAS_MATRICULADAS):
        raise HTTPException(status_code=403)


def _redirect(status: str, *, msg: str | None = None, **extra: str) -> RedirectResponse:
    parts = [f"status={status}"]
    if msg:
        parts.append(f"msg={quote(msg)}")
    for key, value in extra.items():
        parts.append(f"{key}={quote(str(value))}")
    return RedirectResponse(f"/admin/catalogo-asignaturas/?{'&'.join(parts)}", status_code=303)


def _load_workbook(file: UploadFile):
    raw = file.file.read()
    if not raw:
        raise ValueError("Archivo vacío")
    if raw[:2] != b"PK":
        raise ValueError("No es un .xlsx válido")
    return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)


@router.get("/", response_class=HTMLResponse)
def admin_subject_catalog_page(
    request: Request,
    user: dict = Depends(load_user_dep),
    etapa: str = "",
    curso: str = "",
):
    _require_perm(user)
    etapa_f = normalize_catalog_etapa(etapa) if (etapa or "").strip() else None
    curso_f: int | None = None
    raw_curso = (curso or "").strip()
    if raw_curso:
        try:
            curso_f = int(raw_curso)
        except ValueError:
            curso_f = None
    allowed_cursos = CATALOG_COURSE_NUMS.get(etapa_f, (1, 2, 3, 4)) if etapa_f else (1, 2, 3, 4)
    if curso_f is not None and curso_f not in allowed_cursos:
        curso_f = None
    filtered = bool(etapa_f or curso_f)
    meta = get_catalog_meta()
    preview = list_catalog_preview(
        limit=None if filtered else 80,
        etapa=etapa_f,
        curso=curso_f,
    )
    return request.app.state.templates.TemplateResponse(
        "admin/catalogo_asignaturas.html",
        ctx(
            request,
            user=user,
            title="Catálogo de asignaturas",
            portal_shell_title="Catálogo de asignaturas",
            catalog_meta=meta,
            preview_rows=preview,
            column_labels=CATALOG_EXCEL_HEADERS,
            column_fields=CATALOG_FIELD_NAMES,
            filter_etapas=CATALOG_STAGE_KEYS,
            filter_cursos=allowed_cursos,
            selected_etapa=etapa_f or "",
            selected_curso=curso_f,
            catalog_filtered=filtered,
            cursos_por_etapa={k: list(v) for k, v in CATALOG_COURSE_NUMS.items()},
        ),
    )


@router.post("/import")
def admin_subject_catalog_import(
    user: dict = Depends(load_user_dep),
    file: UploadFile = File(...),
):
    _require_perm(user)
    try:
        wb = _load_workbook(file)
    except ValueError as exc:
        _log.warning("Import catálogo asignaturas rechazado: %s", exc)
        return _redirect("error", msg=str(exc))
    except Exception:
        _log.exception("Import catálogo: fallo al leer Excel")
        return _redirect("error", msg="No se pudo leer el archivo Excel.")
    idx_to_field, rows = parse_catalog_workbook_rows(wb)
    if "materia_abrev" not in idx_to_field.values() or "curso_asign" not in idx_to_field.values():
        return _redirect(
            "bad_headers",
            msg="Faltan columnas obligatorias: MATERIA (abrev.) y CURSO_ASIGN.",
        )
    if not rows:
        return _redirect("empty", msg="No hay filas válidas en el Excel.")
    try:
        inserted, skipped = replace_subject_catalog(rows)
    except Exception:
        _log.exception("Import catálogo: fallo al guardar en BD")
        return _redirect("error", msg="Error al guardar el catálogo en la base de datos.")
    return _redirect("imported", rows=str(inserted), skipped=str(skipped))


@router.get("/export")
def admin_subject_catalog_export(user: dict = Depends(load_user_dep)):
    _require_perm(user)
    rows = list_catalog_for_export()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogo"
    ws.append(list(CATALOG_EXCEL_HEADERS))
    for r in rows:
        etapa_key = normalize_catalog_etapa(r.get("etapa")) or ""
        ws.append(
            [
                r.get("materia_abrev") or "",
                r.get("materia") or "",
                r.get("estudio") or "",
                "",
                r.get("curso_asignatura"),
                CATALOG_ETAPA_EXPORT_LABELS.get(etapa_key, r.get("etapa") or ""),
                r.get("departamento") or "",
                r.get("horas") if r.get("horas") is not None else "",
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="catalogo-asignaturas.xlsx"'},
    )
