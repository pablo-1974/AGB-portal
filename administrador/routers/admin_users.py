# routers/admin_users.py
"""
Gestión de usuarios (ADMIN).

Funcionalidades:
- Listar usuarios
- Crear usuario (sin contraseña → primer login)
- Editar nombre, email y rol
- Activar / desactivar usuario
- Resetear contraseña (forzar primer login)

Acceso exclusivo para el rol admin.
Incluye salvaguardas para evitar dejar el sistema sin administradores.
"""

from fastapi import (
    APIRouter,
    Request,
    Form,
    HTTPException,
    UploadFile,
    File,
    Depends,
)
from fastapi.responses import HTMLResponse, RedirectResponse, Response

import io

from utils.local_deps import ensure_local_deps

ensure_local_deps()
try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]

from auth import load_user_dep
from context import ctx
from utils.permissions import has_permission
from utils.enums import PERM_GESTION_USUARIOS, ROLE_INVITADO, ROLES_IMPORTABLES, ROLES_TODOS
from utils.text import normalize_for_sort
from utils.time_madrid import format_madrid

from db.users import (
    get_all_users,
    get_user_by_id,
    get_user_by_email,
    create_user_admin,
    update_user_admin,
    set_user_active,
    reset_user_password,
)
from db.login_security import unlock_user_login

router = APIRouter()

USERS_PER_PAGE = 20


# ----------------------------------------------------------------------
# UTILIDADES
# ----------------------------------------------------------------------

def _parse_page(raw: str | None) -> int:
    text = (raw or "").strip()
    if text.isdigit():
        return max(1, int(text))
    return 1


def _users_list_url(
    *,
    page: int | None = None,
    status: str | None = None,
    edit: int | None = None,
    **extra: str,
) -> str:
    from urllib.parse import urlencode

    params: dict[str, str] = {}
    if page and page > 1:
        params["page"] = str(page)
    if status:
        params["status"] = status
    if edit:
        params["edit"] = str(edit)
    for key, value in extra.items():
        if value:
            params[key] = str(value)
    qs = urlencode(params)
    return f"/admin/users?{qs}" if qs else "/admin/users"

def _count_active_admins() -> int:
    """
    Devuelve el número de administradores activos.
    """
    from db.connection import get_db

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT COUNT(*) AS n
                FROM users
                WHERE role = 'admin'
                  AND active = 1
                """
            )
            row = cur.fetchone()
            return int(row["n"] if row else 0)


def _require_perm(user: dict):
    if not has_permission(user, PERM_GESTION_USUARIOS):
        raise HTTPException(status_code=403)


def _to_bool(value, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    return text in {"1", "si", "sí", "true", "x", "yes", "y"}


def _fields_for_role(
    role: str,
    *,
    status: str,
    titular: object,
    tutor: str,
    departamento: str,
) -> tuple[str, bool, str | None, str | None]:
    """El invitado no es profesorado: sin status docente, tutor, departamento ni titular."""
    if str(role or "").strip().lower() == ROLE_INVITADO:
        return "activo", False, None, None
    return (
        (status or "").strip() or "activo",
        _to_bool(titular, default=True),
        (tutor or "").strip() or None,
        (departamento or "").strip() or None,
    )


# ----------------------------------------------------------------------
# LISTADO DE USUARIOS
# ----------------------------------------------------------------------

@router.get("/admin/users", response_class=HTMLResponse)
def admin_users(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_perm(user)

    all_users = get_all_users()
    total_users = len(all_users)
    page = _parse_page(request.query_params.get("page"))
    total_pages = max(1, (total_users + USERS_PER_PAGE - 1) // USERS_PER_PAGE)
    if page > total_pages:
        page = total_pages
    start = (page - 1) * USERS_PER_PAGE
    users_page = all_users[start:start + USERS_PER_PAGE]
    page_start = start + 1 if total_users else 0
    page_end = min(start + USERS_PER_PAGE, total_users)

    edit_user = None
    raw_edit = (request.query_params.get("edit") or "").strip()
    if raw_edit.isdigit():
        edit_user = get_user_by_id(int(raw_edit))

    return request.app.state.templates.TemplateResponse(
        "admin/users.html",
        ctx(
            request,
            user=user,
            title="Gestión de usuarios",
            users=users_page,
            roles=sorted(ROLES_TODOS, key=normalize_for_sort),
            edit_user=edit_user,
            page=page,
            total_pages=total_pages,
            total_users=total_users,
            users_per_page=USERS_PER_PAGE,
            page_start=page_start,
            page_end=page_end,
        ),
    )


# ----------------------------------------------------------------------
# CREAR USUARIO
# ----------------------------------------------------------------------

@router.post("/admin/users/create")
def admin_users_create(
    request: Request,
    user: dict = Depends(load_user_dep),
    name: str = Form(...),
    email: str = Form(...),
    role: str = Form(...),
    alias: str = Form(""),
    status: str = Form("activo"),
    titular: str = Form("1"),
    tutor: str = Form(""),
    departamento: str = Form(""),
):
    _require_perm(user)

    if role not in ROLES_TODOS:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    status_v, titular_v, tutor_v, dept_v = _fields_for_role(
        role,
        status=status,
        titular=titular,
        tutor=tutor,
        departamento=departamento,
    )
    create_user_admin(
        name=name.strip(),
        email=email.strip(),
        role=role,
        created_by=user["id"],
        alias=alias.strip() or None,
        status=status_v,
        titular=titular_v,
        tutor=tutor_v,
        departamento=dept_v,
    )

    return RedirectResponse("/admin/users?status=created", status_code=303)


# ----------------------------------------------------------------------
# EDITAR USUARIO
# ----------------------------------------------------------------------

@router.post("/admin/users/update/{user_id}")
def admin_users_update(
    request: Request,
    user_id: int,
    user: dict = Depends(load_user_dep),
    name: str = Form(...),
    email: str = Form(...),
    role: str = Form(...),
    alias: str = Form(""),
    status: str = Form("activo"),
    titular: str = Form("1"),
    tutor: str = Form(""),
    departamento: str = Form(""),
    page: str = Form(""),
):
    _require_perm(user)

    if role not in ROLES_TODOS:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    target = get_user_by_id(user_id)
    if not target:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    # Evitar quitar el último admin
    if target["role"] == "admin" and role != "admin":
        if _count_active_admins() <= 1:
            return RedirectResponse("/admin/users?status=error", status_code=303)

    status_v, titular_v, tutor_v, dept_v = _fields_for_role(
        role,
        status=status,
        titular=titular,
        tutor=tutor,
        departamento=departamento,
    )
    update_user_admin(
        user_id=user_id,
        name=name.strip(),
        email=email.strip(),
        role=role,
        alias=alias.strip() or None,
        status=status_v,
        titular=titular_v,
        tutor=tutor_v,
        departamento=dept_v,
        set_departamento=True,
    )

    return RedirectResponse(_users_list_url(status="updated", page=_parse_page(page)), status_code=303)


# ----------------------------------------------------------------------
# ACTIVAR / DESACTIVAR USUARIO
# ----------------------------------------------------------------------

@router.post("/admin/users/toggle/{user_id}")
def admin_users_toggle(
    request: Request,
    user_id: int,
    user: dict = Depends(load_user_dep),
    page: str = Form(""),
):
    _require_perm(user)

    target = get_user_by_id(user_id)
    if not target:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    # Evitar desactivar el último admin activo
    if target["role"] == "admin" and target["active"] == 1:
        if _count_active_admins() <= 1:
            return RedirectResponse("/admin/users?status=error", status_code=303)

    set_user_active(
        user_id=user_id,
        active=not bool(target["active"]),
    )

    return RedirectResponse(
        _users_list_url(status="toggled", page=_parse_page(page)),
        status_code=303,
    )


# ----------------------------------------------------------------------
# RESET DE CONTRASEÑA
# ----------------------------------------------------------------------

@router.post("/admin/users/reset-password/{user_id}")
def admin_users_reset_password(
    request: Request,
    user_id: int,
    user: dict = Depends(load_user_dep),
    page: str = Form(""),
):
    _require_perm(user)

    target = get_user_by_id(user_id)
    if not target:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    reset_user_password(user_id=user_id)

    return RedirectResponse(
        _users_list_url(status="reset", page=_parse_page(page)),
        status_code=303,
    )


@router.post("/admin/users/unlock-login/{user_id}")
def admin_users_unlock_login(
    request: Request,
    user_id: int,
    user: dict = Depends(load_user_dep),
    page: str = Form(""),
):
    _require_perm(user)

    target = get_user_by_id(user_id)
    if not target:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    unlock_user_login(user_id=user_id)
    return RedirectResponse(
        _users_list_url(status="unlocked", page=_parse_page(page)),
        status_code=303,
    )


# ----------------------------------------------------------------------
# IMPORTAR USUARIOS (EXCEL)
# ----------------------------------------------------------------------

@router.post("/admin/users/import")
def admin_users_import(
    user: dict = Depends(load_user_dep),
    file: UploadFile = File(...),
):
    _require_perm(user)

    if not file.filename.lower().endswith(".xlsx"):
        return RedirectResponse("/admin/users?status=error", status_code=303)

    try:
        wb = openpyxl.load_workbook(file.file)
        ws = wb.active
    except Exception:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    headers = [cell.value for cell in ws[1]]
    normalized_headers = [str(h).strip() if h is not None else "" for h in headers]
    required = ["Nombre", "Email", "Rol"]
    if normalized_headers[:3] != required:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    idx = {name: pos for pos, name in enumerate(normalized_headers) if name}

    created = 0
    updated = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name, email, role = row[:3]

        role_v = str(role or "").strip().lower()
        if not email or not role_v or role_v not in ROLES_IMPORTABLES:
            continue

        email = str(email).strip()
        name = name.strip() if name else ""

        alias = row[idx["Alias"]] if "Alias" in idx and idx["Alias"] < len(row) else None
        status = row[idx["Status"]] if "Status" in idx and idx["Status"] < len(row) else None
        titular = row[idx["Titular"]] if "Titular" in idx and idx["Titular"] < len(row) else None
        tutor = row[idx["Tutor"]] if "Tutor" in idx and idx["Tutor"] < len(row) else None
        active = row[idx["Activo"]] if "Activo" in idx and idx["Activo"] < len(row) else None
        force_first_login = (
            row[idx["Primer login pendiente"]]
            if "Primer login pendiente" in idx and idx["Primer login pendiente"] < len(row)
            else None
        )
        departamento = (
            row[idx["Departamento"]]
            if "Departamento" in idx and idx["Departamento"] < len(row)
            else None
        )

        alias_val = str(alias).strip() if alias not in (None, "") else None
        status_val = str(status).strip() if status not in (None, "") else None
        tutor_val = str(tutor).strip() if tutor not in (None, "") else None
        titular_val = _to_bool(titular, default=True)
        active_val = _to_bool(active, default=True)
        force_first_login_val = _to_bool(force_first_login, default=False)
        dept_val = (
            str(departamento).strip()
            if departamento not in (None, "")
            else None
        )

        existing = get_user_by_email(email)
        if existing and str(existing.get("role") or "").strip().lower() == ROLE_INVITADO:
            continue

        if existing:
            upd_kwargs = dict(
                user_id=existing["id"],
                name=name or existing["name"],
                email=email,
                role=role_v,
                alias=alias_val if alias_val is not None else existing.get("alias"),
                status=status_val if status_val is not None else (existing.get("status") or "activo"),
                titular=titular_val if titular is not None else bool(existing.get("titular", True)),
                tutor=tutor_val if tutor_val is not None else existing.get("tutor"),
            )
            if "Departamento" in idx:
                upd_kwargs["departamento"] = dept_val
                upd_kwargs["set_departamento"] = True
            update_user_admin(**upd_kwargs)
            if active is not None:
                set_user_active(user_id=existing["id"], active=active_val)
            if force_first_login is not None and force_first_login_val:
                reset_user_password(user_id=existing["id"])
            updated += 1
        else:
            create_kwargs = dict(
                name=name,
                email=email,
                role=role_v,
                created_by=user["id"],
                alias=alias_val,
                status=status_val or "activo",
                titular=titular_val,
                tutor=tutor_val,
            )
            if "Departamento" in idx:
                create_kwargs["departamento"] = dept_val
            create_user_admin(**create_kwargs)
            if not active_val:
                created_user = get_user_by_email(email)
                if created_user:
                    set_user_active(user_id=created_user["id"], active=False)
            if force_first_login is not None and force_first_login_val:
                created_user = get_user_by_email(email)
                if created_user:
                    reset_user_password(user_id=created_user["id"])
            created += 1

    return RedirectResponse(
        f"/admin/users?status=imported&created={created}&updated={updated}",
        status_code=303,
    )


# ----------------------------------------------------------------------
# EXPORTAR USUARIOS (EXCEL)
# ----------------------------------------------------------------------

def _users_export_headers() -> list[str]:
    return [
        "Nombre",
        "Email",
        "Rol",
        "Alias",
        "Status",
        "Titular",
        "Tutor",
        "Departamento",
        "Activo",
        "Primer login pendiente",
        "Último acceso",
    ]


@router.get("/admin/users/export")
def export_users(
    user: dict = Depends(load_user_dep),
):
    _require_perm(user)

    users = get_all_users()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    ws.append(_users_export_headers())

    for u in users:
        ws.append([
            u["name"],
            u["email"],
            u["role"],
            u.get("alias") or "",
            u.get("status") or "",
            "Sí" if u.get("titular") else "No",
            u.get("tutor") or "",
            u.get("departamento") or "",
            "Sí" if u["active"] == 1 else "No",
            "Sí" if u["must_change_password"] else "No",
            format_madrid(u["last_login_at"]) if u["last_login_at"] else "",
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return Response(
        stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=usuarios.xlsx"
        },
    )
