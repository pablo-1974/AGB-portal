# routers/backup.py

from fastapi import (
    APIRouter, Request, Depends, HTTPException,
    UploadFile, File, Form
)
from fastapi.responses import Response, HTMLResponse, RedirectResponse
from psycopg import sql
from psycopg.types.json import Json
from urllib.parse import urlencode
from datetime import date, datetime, time
from decimal import Decimal
from uuid import UUID
import io
import json
import re

from utils.local_deps import ensure_local_deps

ensure_local_deps()
try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]

from auth import load_user_dep
from db.action_logs import list_action_logs, log_incident_action
from utils.permissions import has_permission
from utils.enums import PERM_BACKUP
from db.connection import get_db
from context import ctx

router = APIRouter()

_EXCEL_INVALID_SHEET_CHARS = set("[]:*?/\\")
_TABLE_NAME_RE = re.compile(r"^[a-z_][a-z0-9_]*$", re.IGNORECASE)
_BACKUP_IMPORT_BLOCKED = frozenset({"users", "alembic_version"})

# Orden sugerido al importar copias completas (padres antes que hijos).
_BACKUP_IMPORT_ORDER: tuple[str, ...] = (
    "school_calendar",
    "groups",
    "departamentos",
    "students",
    "schedule_slots",
    "leaves",
    "absences",
    "enrolled_subjects",
    "incidents",
    "room_reservations",
    "room_reservations_recurring",
    "moscosos_calendar_config",
    "moscosos_reservations",
    "extraescolares",
    "extraescolar_alumnos",
    "extraescolar_acompanantes",
    "funcionamiento_portal_feedback",
    "mantenimiento_feedback",
    "listados_feedback",
    "action_logs",
)

_EXTRAESCOLARES_TABLES = frozenset(
    {"extraescolares", "extraescolar_alumnos", "extraescolar_acompanantes"}
)
_MOSCOSOS_TABLES = frozenset({"moscosos_calendar_config", "moscosos_reservations"})
_EMPTY_PROTECTED_TABLES = frozenset({"users"})

# Datos reales / de catálogo: salen desmarcadas (no se vacían si no se marca a mano).
_EMPTY_KEEP_UNCHECKED = frozenset(
    {
        "users",
        "students",
        "groups",
        "departamentos",
        "school_calendar",
        "school_calendar",
        "enrolled_subject_catalog",
        "enrolled_subjects",
        "enrolled_subjects_imports",
        "competencias_clave",
        "competencias_materia_criterios",
        "competencias_materia_variables",
    }
)

# Qué almacena cada tabla (nombres reales + alias por si Neon usa otra grafía).
_TABLE_STORES: dict[str, str] = {
    "users": "Cuentas del portal: nombre, email, rol, alias, departamento, tutoría, normas aceptadas y acceso.",
    "students": "Alumnado: grupo, nombre, sexo, contactos, CIE, documento, transporte y observaciones.",
    "groups": "Catálogo de grupos y el curso asociado (ESO, Bachillerato, FP…).",
    "school_calendar": "Calendario escolar: inicio/fin de curso, vacaciones, festivos y fin de etapa.",
    "school_calendar": "Calendario escolar: inicio/fin de curso, vacaciones, festivos y fin de etapa.",
    "departamentos": "Departamentos didácticos: abreviatura, nombre y jefatura.",
    "schedule_slots": "Horario semanal de cada profesor: día, franja, tipo (clase/guardia/otros), grupo, aula y materia.",
    "schedule_slots": "Horario semanal de cada profesor: día, franja, tipo (clase/guardia/otros), grupo, aula y materia.",
    "leaves": "Bajas y sustituciones del profesorado: fechas, causa y titular/sustituto.",
    "absences": "Parte diario de ausencias: profesor, fecha, horas y categoría.",
    "action_logs": "Registro de acciones de las apps (quién, qué, cuándo y detalle).",
    "incidents": "Partes de incidencia: profesor, grupo, alumno, fecha, descripción, gravedad y estado.",
    "paa_procedimientos": "Procedimientos PAA (suspensión de asistencia): alumno, grupo, fechas y aviso asociado.",
    "expedientes_disciplinarios": "Expedientes disciplinarios: fechas, medida cautelar, sanción, instructor y avisos.",
    "room_reservations": "Reservas puntuales de aula: grupo, aula, fecha, franja y titular.",
    "room_reservations_recurring": "Reservas recurrentes de aula: día de la semana, franja y periodo de vigencia.",
    "moscosos_calendar_config": "Configuración de días reservables de moscoso y exclusiones del calendario.",
    "moscosos_reservations": "Reservas de moscoso: usuario, fecha, trimestre, plaza y documentación.",
    "extraescolares": "Actividades extraescolares: fecha, nombre, lugar, departamento, responsable y estado.",
    "extraescolar_alumnos": "Alumnado inscrito en cada actividad extraescolar.",
    "extraescolar_acompanantes": "Profesorado acompañante de cada actividad extraescolar.",
    "extraescolar_acompanantes": "Profesorado acompañante de cada actividad extraescolar.",
    "enrolled_subjects_imports": "Cabecera de cada importación de asignaturas matriculadas.",
    "enrolled_subjects": "Filas de matrícula importadas: alumno, materia, curso, grupo y departamento.",
    "enrolled_subject_catalog": "Catálogo de materias: nombre, etapa, curso, departamento, horas y peso.",
    "portal_espacio_visibility": "Visibilidad de cada espacio del portal: visible, en obras u oculto.",
    "portal_published_notices": "Avisos publicados en el portal (alta/baja, sustitución, PAA, expediente, aviso libre).",
    "portal_published_notice_dismissals": "Qué usuario ha ocultado qué aviso del portal.",
    "funcionamiento_portal_feedback": "Buzón de funcionamiento del portal: incidencias o sugerencias.",
    "mantenimiento_feedback": "Buzón de mantenimiento (edificio / informática).",
    "listados_feedback": "Buzón de listados: errores o sugerencias sobre consultas.",
    "portal_feedback": "Buzón único antiguo. Ya no se escribe; se conservan filas migradas si las hay.",
    "competencias_clave": "Las 8 competencias clave LOMLOE y sus descriptores operativos de ESO y Bachillerato.",
    "competencias_materia_criterios": "Criterios de evaluación por materia, competencia específica y descriptores.",
    "competencias_materia_pd_porcentajes": "Peso de cada criterio en la programación didáctica por materia.",
    "competencias_materia_variables": "Cruce criterio × descriptor: peso PD, coeficientes de cálculo y horas.",
    "competencias_pd_edicion": "Si los jefes de departamento tienen bloqueada la edición de porcentajes PD.",
    "competencias_calculo_config": "Opciones globales de cálculo: promedio de descriptores, periodos, pendientes y redondeo.",
    "competencias_fechas_sesion": "Fecha de cada sesión de evaluación por grupo (ordinaria / extraordinaria).",
    "competencias_evaluacion_notas": "Notas por criterio al calificar (sesión ordinaria).",
    "competencias_evaluacion_notas_extra": "Notas por criterio de la sesión extraordinaria de Bachillerato.",
    "competencias_evaluacion_nota_acta": "Nota de acta de la materia por alumno (ordinaria).",
    "competencias_evaluacion_nota_acta_extra": "Nota de acta de la materia en extraordinaria.",
    "competencias_evaluacion_nota_comp": "Nota de competencia de la materia por alumno (ordinaria).",
    "competencias_evaluacion_nota_comp_extra": "Nota de competencia de la materia en extraordinaria.",
    "competencias_sesion_notas": "Notas editadas en la sesión de evaluación (override de materia o competencia).",
    "competencias_do_pesos": "Pesos precalculados de cada cruce descriptor–criterio por materia.",
    "competencias_alumno_materia_do": "Suma ponderada por descriptor dentro de cada materia (ordinaria).",
    "competencias_alumno_materia_do_extra": "Suma ponderada por descriptor dentro de cada materia (extraordinaria).",
    "competencias_alumno_descriptor_notas": "Nota agregada de cada descriptor operativo por alumno (ordinaria).",
    "competencias_alumno_descriptor_notas_extra": "Nota de cada descriptor operativo en extraordinaria.",
    "competencias_alumno_competencia_notas": "Nota de cada competencia clave por alumno (ordinaria).",
    "competencias_alumno_competencia_notas_extra": "Nota de cada competencia clave en extraordinaria.",
    "competencias_bach_ordinaria_acta": "Congelación del acta ordinaria de Bachillerato al pasar a extraordinaria.",
    "schedule_slots": "Horario semanal de cada profesor: día, franja, tipo (clase/guardia/otros), grupo, aula y materia.",
    "school_calendar": "Calendario escolar: inicio/fin de curso, vacaciones, festivos y fin de etapa.",
    "schedule_slots": "Horario semanal de cada profesor: día, franja, tipo (clase/guardia/otros), grupo, aula y materia.",
    "extraescolar_acompanantes": "Profesorado acompañante de cada actividad extraescolar.",
    "enrolled_subjects_imports": "Cabecera de cada importación de asignaturas matriculadas.",
    "enrolled_subjects": "Filas de matrícula importadas: alumno, materia, curso, grupo y departamento.",
    "enrolled_subject_catalog": "Catálogo de materias: nombre, etapa, curso, departamento, horas y peso.",
    "portal_espacio_visibility": "Visibilidad de cada espacio del portal: visible, en obras u oculto.",
    "portal_published_notices": "Avisos publicados en el portal (alta/baja, sustitución, PAA, expediente, aviso libre).",
    "portal_published_notice_dismissals": "Qué usuario ha ocultado qué aviso del portal.",
    "funcionamiento_portal_feedback": "Buzón de funcionamiento del portal: incidencias o sugerencias.",
    "mantenimiento_feedback": "Buzón de mantenimiento (edificio / informática).",
    "listados_feedback": "Buzón de listados: errores o sugerencias sobre consultas.",
    "competencias_clave": "Las 8 competencias clave LOMLOE y sus descriptores operativos de ESO y Bachillerato.",
    "competencias_materia_criterios": "Criterios de evaluación por materia, competencia específica y descriptores.",
    "competencias_materia_pd_porcentajes": "Peso de cada criterio en la programación didáctica por materia.",
    "competencias_materia_variables": "Cruce criterio × descriptor: peso PD, coeficientes de cálculo y horas.",
    "competencias_pd_edicion": "Si los jefes de departamento tienen bloqueada la edición de porcentajes PD.",
    "competencias_calculo_config": "Opciones globales de cálculo: promedio de descriptores, periodos, pendientes y redondeo.",
    "competencias_fechas_sesion": "Fecha de cada sesión de evaluación por grupo (ordinaria / extraordinaria).",
    "competencias_evaluacion_notas": "Notas por criterio al calificar (sesión ordinaria).",
    "competencias_evaluacion_notas_extra": "Notas por criterio de la sesión extraordinaria de Bachillerato.",
    "competencias_evaluacion_nota_acta": "Nota de acta de la materia por alumno (ordinaria).",
    "competencias_evaluacion_nota_acta_extra": "Nota de acta de la materia en extraordinaria.",
    "competencias_evaluacion_nota_comp": "Nota de competencia de la materia por alumno (ordinaria).",
    "competencias_evaluacion_nota_comp_extra": "Nota de competencia de la materia en extraordinaria.",
    "competencias_sesion_notas": "Notas editadas en la sesión de evaluación (override de materia o competencia).",
    "competencias_do_pesos": "Pesos precalculados de cada cruce descriptor–criterio por materia.",
    "competencias_alumno_materia_do": "Suma ponderada por descriptor dentro de cada materia (ordinaria).",
    "competencias_alumno_materia_do_extra": "Suma ponderada por descriptor dentro de cada materia (extraordinaria).",
    "competencias_alumno_descriptor_notas": "Nota agregada de cada descriptor operativo por alumno (ordinaria).",
    "competencias_alumno_descriptor_notas_extra": "Nota de cada descriptor operativo en extraordinaria.",
    "competencias_alumno_competencia_notas": "Nota de cada competencia clave por alumno (ordinaria).",
    "competencias_alumno_competencia_notas_extra": "Nota de cada competencia clave en extraordinaria.",
}


def _stores_for_table(name: str) -> str:
    if name in _TABLE_STORES:
        return _TABLE_STORES[name]
    if name.startswith("competencias_"):
        return "Datos de Evaluación de competencias."
    if name.startswith("portal_"):
        return "Datos del portal (avisos, visibilidad o buzones)."
    if name.startswith("extraescolar"):
        return "Datos de actividades extraescolares."
    if name.startswith("moscosos_"):
        return "Datos de moscosos."
    if name.startswith("enrolled_"):
        return "Datos de matrícula de asignaturas."
    return "Datos de la aplicación."


# Mismos bloques que el inventario PDF (Núcleo, Ausencias, Incidencias…).
_PDF_GROUP_ORDER: tuple[str, ...] = (
    "Núcleo",
    "Ausencias y horarios",
    "Incidencias",
    "Reservas de aulas",
    "Moscosos",
    "Extraescolares",
    "Matrícula",
    "Portal y buzones",
    "Competencias",
    "Otros",
)


def _pdf_group_for_table(table: str) -> str:
    if table in (
        "users",
        "students",
        "groups",
        "departamentos",
        "school_calendar",
        "school_calendar",
    ):
        return "Núcleo"
    if table in (
        "schedule_slots",
        "schedule_slots",
        "leaves",
        "absences",
        "action_logs",
    ) or table.startswith("schedule_"):
        return "Ausencias y horarios"
    if table in ("incidents", "paa_procedimientos", "expedientes_disciplinarios"):
        return "Incidencias"
    if table.startswith("room_reservations") or table.startswith("room_reservations"):
        return "Reservas de aulas"
    if table in _MOSCOSOS_TABLES or table.startswith("moscosos_"):
        return "Moscosos"
    if table in _EXTRAESCOLARES_TABLES or table.startswith("extraescolar"):
        return "Extraescolares"
    if table.startswith("enrolled_"):
        return "Matrícula"
    if table.startswith("competencias_"):
        return "Competencias"
    if (
        table.startswith("portal_published")
        or table.startswith("portal_")
        or table in (
            "funcionamiento_portal_feedback",
            "mantenimiento_feedback",
            "listados_feedback",
            "portal_feedback",
            "portal_espacio_visibility",
            "portal_espacio_visibility",
        )
    ):
        return "Portal y buzones"
    return "Otros"


def _catalog_public_tables() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for name in _list_public_tables():
        protected = name in _EMPTY_PROTECTED_TABLES
        rows.append(
            {
                "name": name,
                "stores": _stores_for_table(name),
                "protected": protected,
                "prechecked": not protected and name not in _EMPTY_KEEP_UNCHECKED,
                "group": _pdf_group_for_table(name),
            }
        )
    return rows


def _catalog_public_tables_grouped() -> list[dict[str, object]]:
    buckets: dict[str, list[dict[str, object]]] = {}
    for row in _catalog_public_tables():
        buckets.setdefault(str(row["group"]), []).append(row)
    out: list[dict[str, object]] = []
    for label in _PDF_GROUP_ORDER:
        tables = buckets.pop(label, None)
        if tables:
            out.append({"label": label, "tables": tables})
    for label in sorted(buckets):
        out.append({"label": label, "tables": buckets[label]})
    return out


def _empty_selected_tables(names: list[str]) -> tuple[list[str], str | None]:
    allowed = set(_list_public_tables())
    selected: list[str] = []
    seen: set[str] = set()
    for raw in names:
        name = (raw or "").strip()
        if name in seen:
            continue
        seen.add(name)
        if name not in allowed or name in _EMPTY_PROTECTED_TABLES:
            continue
        if not _TABLE_NAME_RE.match(name):
            continue
        selected.append(name)
    if not selected:
        return [], "No hay tablas válidas para vaciar."

    rank = {name: idx for idx, name in enumerate(_BACKUP_IMPORT_ORDER)}
    ordered = sorted(selected, key=lambda t: (rank.get(t, 400), t), reverse=True)

    with get_db() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    sql.SQL("TRUNCATE TABLE {} RESTART IDENTITY").format(
                        sql.SQL(", ").join(sql.Identifier(t) for t in ordered)
                    )
                )
                return ordered, None
            except Exception:
                conn.rollback()
            emptied: list[str] = []
            for table in ordered:
                try:
                    cur.execute(
                        sql.SQL("TRUNCATE TABLE {} RESTART IDENTITY").format(
                            sql.Identifier(table)
                        )
                    )
                    emptied.append(table)
                except Exception:
                    conn.rollback()
                    try:
                        cur.execute(
                            sql.SQL("DELETE FROM {}").format(sql.Identifier(table))
                        )
                        emptied.append(table)
                    except Exception:
                        conn.rollback()
                        pending = table
                        extra = f" Vacías: {', '.join(emptied)}." if emptied else ""
                        return emptied, (
                            f"No se pudo vaciar {pending} (dependencias u otra restricción)."
                            " Seleccione también las tablas relacionadas."
                            + extra
                        )
    return ordered, None


def _excel_cell_value(value):
    """Convierte tipos de PostgreSQL no soportados por openpyxl."""
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, time):
        if value.tzinfo is not None:
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, (list, tuple, dict)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def _excel_sheet_name(table_name: str, used: set[str]) -> str:
    base = "".join(c if c not in _EXCEL_INVALID_SHEET_CHARS else "_" for c in table_name)
    base = (base or "tabla")[:31]
    name = base
    n = 2
    while name in used:
        suffix = f"_{n}"
        name = f"{base[: 31 - len(suffix)]}{suffix}"
        n += 1
    used.add(name)
    return name


def _ensure_backup_schemas() -> None:
    """Garantiza tablas de apps integradas antes de listar o exportar."""
    from ausencias.db import ensure_ausencias_schema
    from db.action_logs import ensure_action_logs_schema
    from db.enrolled_subjects import ensure_enrolled_subjects_schema
    from db.extraescolares_schema import ensure_extraescolares_schema
    from db.funcionamiento_portal_feedback import ensure_funcionamiento_portal_feedback_schema
    from db.groups import ensure_groups_schema
    from db.departamentos import ensure_departamentos_schema
    from db.mantenimiento_feedback import ensure_mantenimiento_feedback_schema
    from db.listados_feedback import ensure_listados_feedback_schema
    from db.moscosos_calendar import ensure_moscosos_calendar_schema
    from db.school_calendar import ensure_school_calendar_schema
    from db.moscosos_reservations import ensure_moscosos_reservations_schema
    from db.reservas_access import ensure_reservas_normas_schema
    from db.moscosos_access import ensure_moscosos_normas_schema
    from db.extraescolares_access import ensure_extraescolares_normas_schema
    from db.incidencias_access import ensure_incidencias_normas_schema
    from db.competencias_access import ensure_competencias_normas_schema
    from db.students import ensure_students_schema
    from reservas.db import ensure_reservas_schema

    ensure_groups_schema()
    ensure_departamentos_schema()
    ensure_students_schema()
    ensure_enrolled_subjects_schema()
    ensure_ausencias_schema()
    ensure_reservas_schema()
    ensure_reservas_normas_schema()
    ensure_moscosos_normas_schema()
    ensure_extraescolares_normas_schema()
    ensure_incidencias_normas_schema()
    ensure_competencias_normas_schema()
    ensure_moscosos_calendar_schema()
    ensure_school_calendar_schema()
    ensure_moscosos_reservations_schema()
    ensure_extraescolares_schema()
    ensure_funcionamiento_portal_feedback_schema()
    ensure_mantenimiento_feedback_schema()
    ensure_listados_feedback_schema()
    ensure_action_logs_schema()


def _table_sheet_map(tables: list[str]) -> dict[str, str]:
    used: set[str] = set()
    return {table: _excel_sheet_name(table, used) for table in tables}


def _module_label_for_table(table: str) -> str:
    if table in (
        "users",
        "students",
        "groups",
        "departamentos",
        "school_calendar",
        "school_calendar",
        "enrolled_subjects",
        "enrolled_subjects",
        "enrolled_subjects_imports",
        "enrolled_subject_catalog",
        "enrolled_subject_catalog",
    ) or table.startswith("enrolled_"):
        return "Portal · datos maestros"
    if table in ("incidents", "paa_procedimientos", "expedientes_disciplinarios"):
        return "Incidencias"
    if table in ("absences", "leaves", "schedule_slots", "schedule_slots"):
        return "Ausencias"
    if table.startswith("room_reservations") or table.startswith("room_reservations"):
        return "Reservas"
    if table in _MOSCOSOS_TABLES or table.startswith("moscosos_"):
        return "Moscosos"
    if table in _EXTRAESCOLARES_TABLES or table.startswith("extraescolar"):
        return "Actividades extraescolares"
    if table.startswith("competencias_"):
        return "Evaluación de competencias"
    if table.startswith("portal_published"):
        return "Publicar avisos"
    if table in ("portal_espacio_visibility", "portal_espacio_visibility"):
        return "Portal · visibilidad"
    if table in ("funcionamiento_portal_feedback", "mantenimiento_feedback", "listados_feedback", "portal_feedback"):
        return "Portal · buzones"
    if table == "action_logs":
        return "Registro de acciones"
    return "Otros"


def _group_tables_by_module(tables: list[str]) -> list[dict[str, object]]:
    groups: dict[str, list[str]] = {}
    for table in tables:
        groups.setdefault(_module_label_for_table(table), []).append(table)
    preferred = (
        "Portal · datos maestros",
        "Incidencias",
        "Ausencias",
        "Reservas",
        "Moscosos",
        "Actividades extraescolares",
        "Evaluación de competencias",
        "Publicar avisos",
        "Portal · visibilidad",
        "Portal · buzones",
        "Registro de acciones",
        "Otros",
    )
    out: list[dict[str, object]] = []
    for label in preferred:
        if label in groups:
            out.append({"label": label, "tables": sorted(groups.pop(label))})
    for label in sorted(groups):
        out.append({"label": label, "tables": sorted(groups[label])})
    return out


def _sort_tables_for_import(tables: list[str]) -> list[str]:
    rank = {name: idx for idx, name in enumerate(_BACKUP_IMPORT_ORDER)}
    return sorted(tables, key=lambda t: (rank.get(t, 500), t))


def _table_row_counts(cur, tables: list[str]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for table in tables:
        cur.execute(f'SELECT COUNT(*) AS n FROM "{table}"')
        row = cur.fetchone()
        counts[table] = int(row["n"] if isinstance(row, dict) else row[0])
    return counts


def _fill_info_worksheet(ws, *, user: dict, tables: list[str], counts: dict[str, int]) -> None:
    sheet_map = _table_sheet_map(tables)
    ws.append(["Aplicación", "Portal del centro (gestión integrada)"])
    ws.append(
        [
            "Módulos",
            "Incidencias · Ausencias · Reservas · Moscosos · Extraescolares · Competencias · Avisos · Portal",
        ]
    )
    ws.append(["Fecha backup", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append(["Generado por", user.get("email") or user.get("name") or ""])
    ws.append([])
    ws.append(["Contenido", "Copia completa de todas las tablas públicas"])
    ws.append([])
    ws.append(["Tablas Moscosos", ", ".join(sorted(_MOSCOSOS_TABLES))])
    ws.append(["Tablas Extraescolares", ", ".join(sorted(_EXTRAESCOLARES_TABLES))])
    ws.append([])
    ws.append(["Módulo", "Tabla", "Hoja Excel", "Registros"])
    for table in tables:
        ws.append(
            [
                _module_label_for_table(table),
                table,
                sheet_map.get(table, table),
                counts.get(table, 0),
            ]
        )


def _worksheet_for_table(
    table: str,
    wb: openpyxl.Workbook,
    sheet_map: dict[str, str] | None = None,
):
    if table in wb.sheetnames:
        return wb[table]
    if sheet_map:
        sheet_name = sheet_map.get(table)
        if sheet_name and sheet_name in wb.sheetnames:
            return wb[sheet_name]
    if len(wb.sheetnames) == 1:
        return wb.active
    return None


def _is_full_backup_workbook(wb: openpyxl.Workbook) -> bool:
    if "INFO" in wb.sheetnames:
        return True
    importable = set(_importable_tables())
    sheet_map = _table_sheet_map(_list_public_tables())
    for table in importable:
        if table in wb.sheetnames:
            return True
        sn = sheet_map.get(table)
        if sn and sn in wb.sheetnames:
            return True
    return False


def _import_full_workbook(wb: openpyxl.Workbook) -> int:
    all_tables = _list_public_tables()
    sheet_map = _table_sheet_map(all_tables)
    total = 0
    for table in _sort_tables_for_import(_importable_tables()):
        ws = _worksheet_for_table(table, wb, sheet_map)
        if ws is None:
            continue
        total += _import_table_worksheet(table, ws)
    return total


def _list_public_tables() -> list[str]:
    _ensure_backup_schemas()
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema = 'public'
                  AND table_type = 'BASE TABLE'
                ORDER BY table_name
                """
            )
            return [r["table_name"] for r in cur.fetchall()]


def _importable_tables() -> list[str]:
    return [t for t in _list_public_tables() if t not in _BACKUP_IMPORT_BLOCKED]


def _validate_table_name(table: str) -> str:
    name = (table or "").strip()
    if not _TABLE_NAME_RE.match(name):
        raise HTTPException(status_code=400, detail="Nombre de tabla no válido")
    allowed = _list_public_tables()
    if name not in allowed:
        raise HTTPException(status_code=404, detail="Tabla no encontrada")
    return name


def _table_db_columns(cur, table: str) -> list[str]:
    cur.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s
        ORDER BY ordinal_position
        """,
        (table,),
    )
    return [r["column_name"] for r in cur.fetchall()]


def _coerce_import_value(value, column: str):
    if value is None:
        return None
    if column == "id" and isinstance(value, float) and value == int(value):
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        if s[0] in "[{":
            try:
                return Json(json.loads(s))
            except json.JSONDecodeError:
                return s
        return s
    return value


def _read_sheet_rows(ws) -> tuple[list[str], list[tuple]]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    headers = [h for h in headers if h]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return headers, rows


def _import_incidents_rows(cur, headers: list[str], rows: list[tuple]) -> int:
    imported = 0
    for row in rows:
        data = dict(zip(headers, row))
        if not all([
            data.get("teacher_id"),
            data.get("fecha"),
            data.get("hora"),
            data.get("grupo"),
            data.get("alumno"),
            data.get("descripcion"),
        ]):
            continue
        cur.execute(
            """
            SELECT 1 FROM incidents
            WHERE teacher_id = %s AND fecha = %s AND hora = %s
              AND grupo = %s AND alumno = %s AND descripcion = %s
            """,
            (
                data["teacher_id"],
                data["fecha"],
                data["hora"],
                data["grupo"],
                data["alumno"],
                data["descripcion"],
            ),
        )
        if cur.fetchone():
            continue
        cur.execute(
            """
            INSERT INTO incidents (
                teacher_id, teacher_name, grupo, alumno, fecha, hora, hora_orden,
                descripcion, gravedad_inicial, estado
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                data["teacher_id"],
                data.get("teacher_name"),
                data["grupo"],
                data["alumno"],
                data["fecha"],
                data["hora"],
                data.get("hora_orden"),
                data["descripcion"],
                data.get("gravedad_inicial"),
                data.get("estado"),
            ),
        )
        imported += 1
    return imported


def _import_generic_table(cur, table: str, headers: list[str], rows: list[tuple]) -> int:
    db_cols = set(_table_db_columns(cur, table))
    use_cols = [h for h in headers if h in db_cols]
    if not use_cols:
        return 0
    imported = 0
    has_id = "id" in use_cols
    for row in rows:
        raw = {use_cols[i]: _coerce_import_value(row[i], use_cols[i]) if i < len(row) else None for i in range(len(use_cols))}
        data = {k: v for k, v in raw.items() if v is not None}
        if not data:
            continue
        if has_id and data.get("id") is not None:
            cur.execute(f'SELECT 1 FROM "{table}" WHERE id = %s', (data["id"],))
            if cur.fetchone():
                continue
        cols = list(data.keys())
        placeholders = ", ".join(["%s"] * len(cols))
        col_sql = ", ".join(f'"{c}"' for c in cols)
        try:
            cur.execute(
                f'INSERT INTO "{table}" ({col_sql}) VALUES ({placeholders})',
                [data[c] for c in cols],
            )
            imported += 1
        except Exception:
            continue
    return imported


def _import_table_worksheet(table: str, ws) -> int:
    headers, rows = _read_sheet_rows(ws)
    if not headers:
        return 0
    with get_db() as conn:
        with conn.cursor() as cur:
            if table == "incidents":
                count = _import_incidents_rows(cur, headers, rows)
            else:
                count = _import_generic_table(cur, table, headers, rows)
        conn.commit()
    return count


def _import_table_from_workbook(table: str, wb: openpyxl.Workbook) -> int:
    sheet_map = _table_sheet_map(_list_public_tables())
    ws = _worksheet_for_table(table, wb, sheet_map)
    if ws is None:
        return 0
    return _import_table_worksheet(table, ws)


def _redirect_copies_import(*, table: str | None, imported: int, error: str | None = None) -> RedirectResponse:
    params: dict[str, str] = {}
    if error:
        params["status"] = "error"
        params["msg"] = error
    else:
        params["status"] = "imported"
        params["imported"] = str(imported)
        if table:
            params["table"] = table
    qs = urlencode(params)
    return RedirectResponse(f"/admin/backup/copies?{qs}", status_code=303)


def _fill_worksheet_from_table(ws, cur, table: str) -> None:
    cur.execute(f'SELECT * FROM "{table}"')
    columns = _column_names(cur)
    if columns:
        ws.append(columns)
    for row in cur.fetchall():
        ws.append([_excel_cell_value(v) for v in row.values()])


def _xlsx_response(wb: openpyxl.Workbook, filename: str) -> Response:
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return Response(
        stream.read(),
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _column_names(cursor) -> list[str]:
    if not cursor.description:
        return []
    names: list[str] = []
    for col in cursor.description:
        if isinstance(col, (list, tuple)):
            names.append(str(col[0]))
        elif hasattr(col, "name"):
            names.append(str(col.name))
        else:
            names.append(str(col))
    return names


# ======================================================
# PÁGINA BACKUP (UI)
# ======================================================
def _require_backup_perm(user: dict) -> None:
    if not has_permission(user, PERM_BACKUP):
        raise HTTPException(status_code=403)


@router.get("/admin/backup", response_class=HTMLResponse)
def backup_hub_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    return request.app.state.templates.TemplateResponse(
        "admin/backup_hub.html",
        ctx(request, user=user, title="Backup"),
    )


@router.get("/admin/backup/copies", response_class=HTMLResponse)
def backup_copies_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    tables = _list_public_tables()
    return request.app.state.templates.TemplateResponse(
        "admin/backup_copies.html",
        ctx(
            request,
            user=user,
            title="Copias de seguridad",
            backup_tables=tables,
            backup_import_tables=_importable_tables(),
            backup_table_groups=_group_tables_by_module(tables),
        ),
    )


@router.get("/admin/backup/tablas", response_class=HTMLResponse)
def backup_tablas_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    return request.app.state.templates.TemplateResponse(
        "admin/backup_tablas.html",
        ctx(
            request,
            user=user,
            title="Tablas · Backup",
            table_groups=_catalog_public_tables_grouped(),
        ),
    )


@router.post("/admin/backup/tablas/vaciar")
def backup_tablas_vaciar(
    user: dict = Depends(load_user_dep),
    tables: list[str] = Form(default=[]),
):
    _require_backup_perm(user)
    emptied, error = _empty_selected_tables(tables)
    if error:
        qs = urlencode({"status": "error", "msg": error})
        return RedirectResponse(f"/admin/backup/tablas?{qs}", status_code=303)
    qs = urlencode({"status": "emptied", "n": str(len(emptied))})
    return RedirectResponse(f"/admin/backup/tablas?{qs}", status_code=303)


@router.get("/admin/backup/registro", response_class=HTMLResponse)
def backup_registro_hub_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    return request.app.state.templates.TemplateResponse(
        "admin/backup_registro_hub.html",
        ctx(request, user=user, title="Registro · Backup"),
    )


@router.get("/admin/backup/registro/ausencias-actions", response_class=HTMLResponse)
def backup_registro_ausencias_actions_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    return request.app.state.templates.TemplateResponse(
        "admin/backup_ausencias_actions.html",
        ctx(
            request,
            user=user,
            title="Registro de acciones (Ausencias) · Backup",
            logs=list_action_logs(limit=300, module="ausencias"),
        ),
    )


@router.get("/admin/backup/registro/incidencias", response_class=HTMLResponse)
def backup_registro_incidencias_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    return request.app.state.templates.TemplateResponse(
        "admin/backup_registro_incidencias.html",
        ctx(
            request,
            user=user,
            title="Registro de acciones (Incidencias) · Backup",
            logs=list_action_logs(limit=300, module="incidencias"),
        ),
    )


@router.get("/admin/backup/registro/reservas", response_class=HTMLResponse)
def backup_registro_reservas_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    return request.app.state.templates.TemplateResponse(
        "admin/backup_registro_reservas.html",
        ctx(
            request,
            user=user,
            title="Registro de acciones (Reservas) · Backup",
            logs=list_action_logs(limit=300, module="reservas"),
        ),
    )


@router.get("/admin/backup/registro/moscosos", response_class=HTMLResponse)
def backup_registro_moscosos_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    return request.app.state.templates.TemplateResponse(
        "admin/backup_registro_moscosos.html",
        ctx(
            request,
            user=user,
            title="Registro de acciones (Moscosos) · Backup",
            logs=list_action_logs(limit=300, module="moscosos"),
        ),
    )


@router.get("/admin/backup/registro/extraescolares", response_class=HTMLResponse)
def backup_registro_extraescolares_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    return request.app.state.templates.TemplateResponse(
        "admin/backup_registro_extraescolares.html",
        ctx(
            request,
            user=user,
            title="Registro de acciones (Extraescolares) · Backup",
            logs=list_action_logs(limit=300, module="extraescolares"),
        ),
    )


@router.get("/admin/backup/registro/publicar-avisos", response_class=HTMLResponse)
def backup_registro_publicar_avisos_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    return request.app.state.templates.TemplateResponse(
        "admin/backup_registro_publicar_avisos.html",
        ctx(
            request,
            user=user,
            title="Registro de acciones (Publicar avisos) · Backup",
            logs=list_action_logs(limit=300, module="publicar_avisos"),
        ),
    )


@router.get("/admin/backup/registro/competencias", response_class=HTMLResponse)
def backup_registro_competencias_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    return request.app.state.templates.TemplateResponse(
        "admin/backup_registro_competencias.html",
        ctx(
            request,
            user=user,
            title="Registro de acciones (Evaluación de competencias) · Backup",
            logs=list_action_logs(limit=300, module="competencias"),
        ),
    )

# ======================================================
# DESCARGA BACKUP
# ======================================================
@router.get("/admin/backup/download")
def backup_download(
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)

    wb = openpyxl.Workbook()
    tables = _list_public_tables()

    ws_info = wb.active
    ws_info.title = "INFO"

    used_sheet_names: set[str] = set()
    with get_db() as conn:
        with conn.cursor() as cur:
            counts = _table_row_counts(cur, tables)
            _fill_info_worksheet(ws_info, user=user, tables=tables, counts=counts)
            for table in tables:
                sheet_title = _excel_sheet_name(table, used_sheet_names)
                ws = wb.create_sheet(title=sheet_title)
                _fill_worksheet_from_table(ws, cur, table)

    filename = f"campus_backup_{datetime.now():%Y%m%d_%H%M}.xlsx"
    return _xlsx_response(wb, filename)


@router.get("/admin/backup/download/{table_name}")
def backup_download_table(
    table_name: str,
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    table = _validate_table_name(table_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _excel_sheet_name(table, set())

    with get_db() as conn:
        with conn.cursor() as cur:
            _fill_worksheet_from_table(ws, cur, table)

    filename = f"backup_{table}_{datetime.now():%Y%m%d_%H%M}.xlsx"
    return _xlsx_response(wb, filename)

# ======================================================
# SUBIDA BACKUP (IMPORTACIÓN INCREMENTAL)
# ======================================================
@router.post("/admin/backup/upload")
def backup_upload(
    file: UploadFile = File(...),
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return _redirect_copies_import(table=None, imported=0, error="Formato no válido")

    try:
        wb = openpyxl.load_workbook(file.file)
    except Exception:
        return _redirect_copies_import(table=None, imported=0, error="No se pudo leer el Excel")

    if _is_full_backup_workbook(wb):
        try:
            imported_count = _import_full_workbook(wb)
        except Exception:
            return _redirect_copies_import(
                table=None,
                imported=0,
                error="Error al importar la copia completa",
            )
    elif "incidents" in wb.sheetnames:
        imported_count = _import_table_from_workbook("incidents", wb)
    else:
        return _redirect_copies_import(
            table=None,
            imported=0,
            error="No se reconoce el formato del Excel (use una copia generada aquí)",
        )

    if imported_count and "incidents" in wb.sheetnames:
        log_incident_action(
            user_id=user.get("id"),
            action="incident_backup_import",
            detail=f"Importación backup: {imported_count} registro(s) nuevo(s)",
        )

    return _redirect_copies_import(table=None, imported=imported_count)


@router.post("/admin/backup/upload/{table_name}")
def backup_upload_table(
    table_name: str,
    file: UploadFile = File(...),
    user: dict = Depends(load_user_dep),
):
    _require_backup_perm(user)
    table = _validate_table_name(table_name)

    if table in _BACKUP_IMPORT_BLOCKED:
        return _redirect_copies_import(
            table=table,
            imported=0,
            error=f"La tabla {table} no se importa aquí",
        )

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return _redirect_copies_import(table=table, imported=0, error="Formato no válido")

    try:
        wb = openpyxl.load_workbook(file.file)
    except Exception:
        return _redirect_copies_import(table=table, imported=0, error="No se pudo leer el Excel")

    try:
        imported_count = _import_table_from_workbook(table, wb)
    except Exception:
        return _redirect_copies_import(table=table, imported=0, error="Error al importar la tabla")

    if table == "incidents" and imported_count:
        log_incident_action(
            user_id=user.get("id"),
            action="incident_backup_import",
            detail=f"Importación tabla incidents: {imported_count} nueva(s)",
        )

    return _redirect_copies_import(table=table, imported=imported_count)
