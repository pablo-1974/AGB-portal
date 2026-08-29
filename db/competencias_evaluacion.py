# -*- coding: utf-8 -*-
# DISK_FLUSH_20260822_1226 — Comodo off; regrabado a C:\dev
"""Calificaciones por alumno x criterio (Evaluar competencias)."""

from __future__ import annotations

import io
import re
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any
import sys
from pathlib import Path

_pydeps = Path(__file__).resolve().parents[1] / "pydeps"
if _pydeps.is_dir():
    _p = str(_pydeps)
    if _p not in sys.path:
        sys.path.insert(0, _p)
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None  # type: ignore[assignment]
    Alignment = Font = PatternFill = get_column_letter = None  # type: ignore[assignment]

from db.competencias_materia_criterios import list_criterios_materia
from db.connection import get_db
from db.enrolled_subject_catalog import (
    competencias_materia_group_key,
    ensure_subject_catalog_schema,
)
from db.enrolled_subjects import (
    CARACTERISTICA_MATERIA_PENDIENTE,
    ensure_enrolled_subjects_schema,
)
from db.students import get_students_by_group
from utils.text import normalize_for_sort, sql_alumno_key

TABLE = "competencias_evaluacion_notas"
TABLE_ACTA = "competencias_evaluacion_nota_acta"
TABLE_EXTRA = "competencias_evaluacion_notas_extra"
TABLE_ACTA_EXTRA = "competencias_evaluacion_nota_acta_extra"
TABLE_COMP = "competencias_evaluacion_nota_comp"
TABLE_COMP_EXTRA = "competencias_evaluacion_nota_comp_extra"
META_SHEET = "_meta"
DATA_SHEET = "Calificaciones"

_schema_ready = False


def ensure_competencias_evaluacion_schema() -> None:
    global _schema_ready
    if _schema_ready:
        return
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS {TABLE} (
                    etapa TEXT NOT NULL,
                    curso_asignatura SMALLINT NOT NULL,
                    materia_key TEXT NOT NULL,
                    grupo TEXT NOT NULL,
                    alumno TEXT NOT NULL,
                    criterio TEXT NOT NULL,
                    nota NUMERIC(8, 4),
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    updated_by INTEGER,
                    PRIMARY KEY (
                        etapa, curso_asignatura, materia_key, grupo, alumno, criterio
                    )
                )
                """
            )
            cur.execute(
                f"""
                CREATE INDEX IF NOT EXISTS idx_cen_grupo_materia
                ON {TABLE} (etapa, curso_asignatura, materia_key, grupo)
                """
            )
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS {TABLE_ACTA} (
                    etapa TEXT NOT NULL,
                    curso_asignatura SMALLINT NOT NULL,
                    materia_key TEXT NOT NULL,
                    grupo TEXT NOT NULL,
                    alumno TEXT NOT NULL,
                    nota NUMERIC(8, 4) NOT NULL,
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    updated_by INTEGER,
                    PRIMARY KEY (
                        etapa, curso_asignatura, materia_key, grupo, alumno
                    )
                )
                """
            )
            cur.execute(
                f"""
                CREATE INDEX IF NOT EXISTS idx_cena_grupo_materia
                ON {TABLE_ACTA} (etapa, curso_asignatura, materia_key, grupo)
                """
            )
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS {TABLE_EXTRA} (
                    etapa TEXT NOT NULL,
                    curso_asignatura SMALLINT NOT NULL,
                    materia_key TEXT NOT NULL,
                    grupo TEXT NOT NULL,
                    alumno TEXT NOT NULL,
                    criterio TEXT NOT NULL,
                    nota NUMERIC(8, 4),
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    updated_by INTEGER,
                    PRIMARY KEY (
                        etapa, curso_asignatura, materia_key, grupo, alumno, criterio
                    )
                )
                """
            )
            cur.execute(
                f"""
                CREATE INDEX IF NOT EXISTS idx_cene_grupo_materia
                ON {TABLE_EXTRA} (etapa, curso_asignatura, materia_key, grupo)
                """
            )
            cur.execute(
                f"""
                CREATE TABLE IF NOT EXISTS {TABLE_ACTA_EXTRA} (
                    etapa TEXT NOT NULL,
                    curso_asignatura SMALLINT NOT NULL,
                    materia_key TEXT NOT NULL,
                    grupo TEXT NOT NULL,
                    alumno TEXT NOT NULL,
                    nota NUMERIC(8, 4) NOT NULL,
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    updated_by INTEGER,
                    PRIMARY KEY (
                        etapa, curso_asignatura, materia_key, grupo, alumno
                    )
                )
                """
            )
            cur.execute(
                f"""
                CREATE INDEX IF NOT EXISTS idx_cenae_grupo_materia
                ON {TABLE_ACTA_EXTRA} (etapa, curso_asignatura, materia_key, grupo)
                """
            )
            for tbl_comp, idx_name in (
                (TABLE_COMP, "idx_cenc_grupo_materia"),
                (TABLE_COMP_EXTRA, "idx_cence_grupo_materia"),
            ):
                cur.execute(
                    f"""
                    CREATE TABLE IF NOT EXISTS {tbl_comp} (
                        etapa TEXT NOT NULL,
                        curso_asignatura SMALLINT NOT NULL,
                        materia_key TEXT NOT NULL,
                        grupo TEXT NOT NULL,
                        alumno TEXT NOT NULL,
                        nota NUMERIC(8, 4) NOT NULL,
                        updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        updated_by INTEGER,
                        PRIMARY KEY (
                            etapa, curso_asignatura, materia_key, grupo, alumno
                        )
                    )
                    """
                )
                cur.execute(
                    f"""
                    CREATE INDEX IF NOT EXISTS {idx_name}
                    ON {tbl_comp} (etapa, curso_asignatura, materia_key, grupo)
                    """
                )
    _schema_ready = True


def _sesion_es_extraordinaria(sesion: str | None) -> bool:
    return (sesion or "").strip().lower() == "extraordinaria"


def _table_notas(sesion: str | None) -> str:
    return TABLE_EXTRA if _sesion_es_extraordinaria(sesion) else TABLE


def _table_acta(sesion: str | None) -> str:
    return TABLE_ACTA_EXTRA if _sesion_es_extraordinaria(sesion) else TABLE_ACTA


def _table_comp(sesion: str | None) -> str:
    return TABLE_COMP_EXTRA if _sesion_es_extraordinaria(sesion) else TABLE_COMP


def _hechas_from_rows(rows) -> set[tuple[str, int, str]]:
    out: set[tuple[str, int, str]] = set()
    for r in rows:
        etapa = str(r["etapa"] or "").strip().lower()
        key = str(r["materia_key"] or "").strip()
        if not etapa or not key:
            continue
        try:
            curso = int(r["curso_asignatura"])
        except (TypeError, ValueError):
            continue
        out.add((etapa, curso, key))
        fam = competencias_materia_group_key(key) or key
        out.add((etapa, curso, fam))
    return out


def _norm_alumno(nombre: str) -> str:
    from utils.text import normalize_alumno_key

    return normalize_alumno_key(nombre)


def mapas_notas_grupo(
    grupo: str,
    *,
    sesion: str | None = None,
) -> tuple[
    dict[tuple[str, int, str], dict[str, set[str]]],
    dict[tuple[str, int, str], set[str]],
]:
    """Notas de competencias y acta del grupo, indexadas por (etapa, curso, key).

    competencias: alumno -> set de criterios con nota.
    acta: set de alumnos con nota_acta.

    Si ``sesion=extraordinaria``, lee solo las tablas de calificaciones extraordinarias.
    """
    ensure_competencias_evaluacion_schema()
    nombre = (grupo or "").strip()
    comp: dict[tuple[str, int, str], dict[str, set[str]]] = {}
    acta: dict[tuple[str, int, str], set[str]] = {}
    if not nombre:
        return comp, acta

    tbl = _table_notas(sesion)
    tbl_acta = _table_acta(sesion)

    def _add_key(dest, etapa: str, curso: int, key: str):
        fam = competencias_materia_group_key(key) or key
        keys = [(etapa, curso, key)]
        if fam != key:
            keys.append((etapa, curso, fam))
        return keys

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT etapa, curso_asignatura, materia_key, alumno, criterio
                FROM {tbl}
                WHERE LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                  AND nota IS NOT NULL
                """,
                (nombre,),
            )
            for r in cur.fetchall():
                etapa = str(r["etapa"] or "").strip().lower()
                key = str(r["materia_key"] or "").strip()
                al = _norm_alumno(str(r["alumno"] or ""))
                crit = str(r["criterio"] or "").strip()
                if not etapa or not key or not al or not crit:
                    continue
                try:
                    curso = int(r["curso_asignatura"])
                except (TypeError, ValueError):
                    continue
                for gkey in _add_key(comp, etapa, curso, key):
                    comp.setdefault(gkey, {}).setdefault(al, set()).add(crit)

            cur.execute(
                f"""
                SELECT etapa, curso_asignatura, materia_key, alumno
                FROM {tbl_acta}
                WHERE LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                  AND nota IS NOT NULL
                """,
                (nombre,),
            )
            for r in cur.fetchall():
                etapa = str(r["etapa"] or "").strip().lower()
                key = str(r["materia_key"] or "").strip()
                al = _norm_alumno(str(r["alumno"] or ""))
                if not etapa or not key or not al:
                    continue
                try:
                    curso = int(r["curso_asignatura"])
                except (TypeError, ValueError):
                    continue
                for gkey in _add_key(acta, etapa, curso, key):
                    acta.setdefault(gkey, set()).add(al)
    return comp, acta


def list_notas_acta(
    *,
    etapa: str,
    curso_asignatura: int,
    materia_key: str,
    grupo: str,
    sesion: str | None = None,
) -> dict[str, Decimal]:
    """Mapa alumno -> nota_acta."""
    ensure_competencias_evaluacion_schema()
    key = competencias_materia_group_key(materia_key) or (materia_key or "").strip()
    tbl = _table_acta(sesion)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT alumno, nota
                FROM {tbl}
                WHERE etapa = %s
                  AND curso_asignatura = %s
                  AND materia_key = %s
                  AND LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                """,
                (
                    (etapa or "").strip().lower(),
                    int(curso_asignatura),
                    key,
                    (grupo or "").strip(),
                ),
            )
            out: dict[str, Decimal] = {}
            for r in cur.fetchall():
                al = str(r["alumno"] or "").strip()
                if not al or r.get("nota") is None:
                    continue
                out[al] = Decimal(str(r["nota"]))
            return out


def mapa_notas_acta_valores(
    grupo: str,
    *,
    sesion: str | None = None,
) -> dict[tuple[str, int, str], dict[str, Decimal]]:
    """nota_acta del grupo: (etapa, curso, key) -> {alumno_norm: nota}."""
    ensure_competencias_evaluacion_schema()
    nombre = (grupo or "").strip()
    out: dict[tuple[str, int, str], dict[str, Decimal]] = {}
    if not nombre:
        return out
    tbl = _table_acta(sesion)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT etapa, curso_asignatura, materia_key, alumno, nota
                FROM {tbl}
                WHERE LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                  AND nota IS NOT NULL
                """,
                (nombre,),
            )
            for r in cur.fetchall():
                etapa = str(r["etapa"] or "").strip().lower()
                key = str(r["materia_key"] or "").strip()
                al = _norm_alumno(str(r["alumno"] or ""))
                if not etapa or not key or not al or r.get("nota") is None:
                    continue
                try:
                    curso = int(r["curso_asignatura"])
                except (TypeError, ValueError):
                    continue
                nota = Decimal(str(r["nota"]))
                fam = competencias_materia_group_key(key) or key
                for gkey in ((etapa, curso, key), (etapa, curso, fam)):
                    out.setdefault(gkey, {})[al] = nota
    return out


def mapa_notas_criterio_valores(
    grupo: str,
    *,
    sesion: str | None = None,
) -> dict[tuple[str, int, str], dict[str, dict[str, Decimal]]]:
    """Notas por criterio del grupo: (etapa, curso, key) -> {alumno_norm: {criterio: nota}}."""
    ensure_competencias_evaluacion_schema()
    nombre = (grupo or "").strip()
    out: dict[tuple[str, int, str], dict[str, dict[str, Decimal]]] = {}
    if not nombre:
        return out
    tbl = _table_notas(sesion)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT etapa, curso_asignatura, materia_key, alumno, criterio, nota
                FROM {tbl}
                WHERE LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                  AND nota IS NOT NULL
                """,
                (nombre,),
            )
            for r in cur.fetchall():
                etapa = str(r["etapa"] or "").strip().lower()
                key = str(r["materia_key"] or "").strip()
                al = _norm_alumno(str(r["alumno"] or ""))
                crit = str(r["criterio"] or "").strip()
                if not etapa or not key or not al or not crit or r.get("nota") is None:
                    continue
                try:
                    curso = int(r["curso_asignatura"])
                except (TypeError, ValueError):
                    continue
                nota = Decimal(str(r["nota"]))
                fam = competencias_materia_group_key(key) or key
                for gkey in ((etapa, curso, key), (etapa, curso, fam)):
                    out.setdefault(gkey, {}).setdefault(al, {})[crit] = nota
    return out


def mapa_notas_comp_valores(
    grupo: str,
    *,
    sesion: str | None = None,
) -> dict[tuple[str, int, str], dict[str, Decimal]]:
    """nota_comp del grupo: (etapa, curso, key) -> {alumno_norm: nota}."""
    ensure_competencias_evaluacion_schema()
    nombre = (grupo or "").strip()
    out: dict[tuple[str, int, str], dict[str, Decimal]] = {}
    if not nombre:
        return out
    tbl = _table_comp(sesion)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT etapa, curso_asignatura, materia_key, alumno, nota
                FROM {tbl}
                WHERE LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                  AND nota IS NOT NULL
                """,
                (nombre,),
            )
            for r in cur.fetchall():
                etapa = str(r["etapa"] or "").strip().lower()
                key = str(r["materia_key"] or "").strip()
                al = _norm_alumno(str(r["alumno"] or ""))
                if not etapa or not key or not al or r.get("nota") is None:
                    continue
                try:
                    curso = int(r["curso_asignatura"])
                except (TypeError, ValueError):
                    continue
                nota = Decimal(str(r["nota"]))
                fam = competencias_materia_group_key(key) or key
                for gkey in ((etapa, curso, key), (etapa, curso, fam)):
                    out.setdefault(gkey, {})[al] = nota
    return out


def replace_notas_comp(
    *,
    etapa: str,
    curso_asignatura: int,
    materia_key: str,
    grupo: str,
    notas: dict[str, Decimal | None],
    updated_by: int | None = None,
    sesion: str | None = None,
) -> None:
    """Sustituye nota_comp del grupo/materia (valores ya calculados)."""
    ensure_competencias_evaluacion_schema()
    etapa_v = (etapa or "").strip().lower()
    curso = int(curso_asignatura)
    key = competencias_materia_group_key(materia_key) or (materia_key or "").strip()
    grupo_v = (grupo or "").strip()
    if not etapa_v or not key or not grupo_v:
        return
    tbl = _table_comp(sesion)
    rows_ins: list[tuple] = []
    for alumno, nota in notas.items():
        al = (alumno or "").strip()
        if not al or nota is None:
            continue
        rows_ins.append((etapa_v, curso, key, grupo_v, al, nota, updated_by))
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                DELETE FROM {tbl}
                WHERE etapa = %s
                  AND curso_asignatura = %s
                  AND materia_key = %s
                  AND LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                """,
                (etapa_v, curso, key, grupo_v),
            )
            if rows_ins:
                cur.executemany(
                    f"""
                    INSERT INTO {tbl} (
                        etapa, curso_asignatura, materia_key, grupo,
                        alumno, nota, updated_at, updated_by
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, NOW(), %s)
                    """,
                    rows_ins,
                )


def grupo_tiene_notas_criterio(grupo: str, *, sesion: str | None = None) -> bool:
    ensure_competencias_evaluacion_schema()
    nombre = (grupo or "").strip()
    if not nombre:
        return False
    tbl = _table_notas(sesion)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT 1 FROM {tbl}
                WHERE LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                  AND nota IS NOT NULL
                LIMIT 1
                """,
                (nombre,),
            )
            return cur.fetchone() is not None


def grupo_tiene_notas_comp(grupo: str, *, sesion: str | None = None) -> bool:
    ensure_competencias_evaluacion_schema()
    nombre = (grupo or "").strip()
    if not nombre:
        return False
    tbl = _table_comp(sesion)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT 1 FROM {tbl}
                WHERE LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                  AND nota IS NOT NULL
                LIMIT 1
                """,
                (nombre,),
            )
            return cur.fetchone() is not None


def ensure_notas_comp_grupo(grupo: str) -> int:
    """Rellena nota_comp si hay criterios sin media (por materia)."""
    nombre = (grupo or "").strip()
    if not nombre:
        return 0
    ensure_competencias_evaluacion_schema()

    def _keys(tbl: str) -> set[tuple[str, int, str]]:
        out: set[tuple[str, int, str]] = set()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT DISTINCT etapa, curso_asignatura, materia_key
                    FROM {tbl}
                    WHERE LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                      AND nota IS NOT NULL
                    """,
                    (nombre,),
                )
                for r in cur.fetchall():
                    etapa = str(r.get("etapa") or "").strip().lower()
                    key = str(r.get("materia_key") or "").strip()
                    if not etapa or not key:
                        continue
                    try:
                        curso = int(r["curso_asignatura"])
                    except (TypeError, ValueError):
                        continue
                    fam = competencias_materia_group_key(key) or key
                    out.add((etapa, curso, fam))
        return out

    need = False
    for sesion in (None, "extraordinaria"):
        crit = _keys(_table_notas(sesion))
        if not crit:
            continue
        comp = _keys(_table_comp(sesion))
        if crit - comp:
            need = True
            break
    if not need:
        return 0
    return refresh_notas_comp_grupo(nombre)


def refresh_notas_comp_grupo(grupo: str) -> int:
    # Recalcula nota_comp del grupo (ordinaria/extra). No al abrir pantallas.
    from db.competencias_pd_porcentajes import list_porcentajes_materia

    ensure_competencias_evaluacion_schema()
    nombre = (grupo or "").strip()
    if not nombre:
        return 0
    updated = 0
    for sesion in (None, "extraordinaria"):
        tbl = _table_notas(sesion)
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT etapa, curso_asignatura, materia_key, alumno, criterio, nota
                    FROM {tbl}
                    WHERE LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                      AND nota IS NOT NULL
                    """,
                    (nombre,),
                )
                rows = list(cur.fetchall())
        if not rows:
            if sesion == "extraordinaria":
                replace_notas_comp_clear_grupo(nombre, sesion=sesion)
            continue

        # (etapa, curso, fam) → alumno_norm → criterio → nota
        by_mat: dict[tuple[str, int, str], dict[str, dict[str, Decimal]]] = {}
        nombre_por_al: dict[str, str] = {}
        for r in rows:
            etapa = str(r.get("etapa") or "").strip().lower()
            key_raw = str(r.get("materia_key") or "").strip()
            al_raw = str(r.get("alumno") or "").strip()
            al = _norm_alumno(al_raw)
            cr = str(r.get("criterio") or "").strip()
            if not etapa or not key_raw or not al or not cr or r.get("nota") is None:
                continue
            try:
                curso = int(r["curso_asignatura"])
            except (TypeError, ValueError):
                continue
            fam = competencias_materia_group_key(key_raw) or key_raw
            by_mat.setdefault((etapa, curso, fam), {}).setdefault(al, {})[cr] = Decimal(
                str(r["nota"])
            )
            # Preferir forma sin espacio antes de coma para persistir.
            prev = nombre_por_al.get(al)
            if prev is None or (" ," in prev and " ," not in al_raw):
                nombre_por_al[al] = al_raw

        ppd_cache: dict[tuple[str, int, str, str], dict[str, Decimal]] = {}
        for (etapa, curso, fam), por_al in by_mat.items():
            ck = (etapa, curso, fam, sesion or "")
            if ck not in ppd_cache:
                ppd_cache[ck] = list_porcentajes_materia(
                    etapa=etapa,
                    curso_asignatura=curso,
                    materia_key=fam,
                    sesion=sesion,
                    pendiente=False,
                )
            ppd_map = ppd_cache[ck]
            replace_notas_comp(
                etapa=etapa,
                curso_asignatura=curso,
                materia_key=fam,
                grupo=nombre,
                notas={
                    nombre_por_al.get(al, al): compute_nota_comp(por_crit, ppd_map)
                    for al, por_crit in por_al.items()
                },
                sesion=sesion,
            )
            updated += 1
    return updated


def replace_notas_comp_clear_grupo(grupo: str, *, sesion: str | None = None) -> None:
    ensure_competencias_evaluacion_schema()
    nombre = (grupo or "").strip()
    if not nombre:
        return
    tbl = _table_comp(sesion)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                DELETE FROM {tbl}
                WHERE LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                """,
                (nombre,),
            )


def replace_notas_acta(
    *,
    etapa: str,
    curso_asignatura: int,
    materia_key: str,
    grupo: str,
    notas: dict[str, Decimal | None],
    updated_by: int | None = None,
    sesion: str | None = None,
) -> None:
    """Sustituye las nota_acta del grupo/materia."""
    ensure_competencias_evaluacion_schema()
    etapa_v = (etapa or "").strip().lower()
    curso = int(curso_asignatura)
    key = competencias_materia_group_key(materia_key) or (materia_key or "").strip()
    grupo_v = (grupo or "").strip()
    if not etapa_v or not key or not grupo_v:
        raise ValueError("Datos de evaluacion incompletos")

    tbl = _table_acta(sesion)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                DELETE FROM {tbl}
                WHERE etapa = %s
                  AND curso_asignatura = %s
                  AND materia_key = %s
                  AND LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                """,
                (etapa_v, curso, key, grupo_v),
            )
            for alumno, nota in notas.items():
                al = (alumno or "").strip()
                if not al or nota is None:
                    continue
                cur.execute(
                    f"""
                    INSERT INTO {tbl} (
                        etapa, curso_asignatura, materia_key, grupo,
                        alumno, nota, updated_at, updated_by
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, NOW(), %s)
                    """,
                    (etapa_v, curso, key, grupo_v, al, nota, updated_by),
                )


def set_materias_con_notas_grupo(grupo: str) -> set[tuple[str, int, str]]:
    """Pares (etapa, curso, materia_key) con al menos una nota en el grupo."""
    ensure_competencias_evaluacion_schema()
    nombre = (grupo or "").strip()
    if not nombre:
        return set()
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT DISTINCT etapa, curso_asignatura, materia_key
                FROM {TABLE}
                WHERE LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                  AND nota IS NOT NULL
                """,
                (nombre,),
            )
            out: set[tuple[str, int, str]] = set()
            for r in cur.fetchall():
                etapa = str(r["etapa"] or "").strip().lower()
                key = str(r["materia_key"] or "").strip()
                if not etapa or not key:
                    continue
                try:
                    curso = int(r["curso_asignatura"])
                except (TypeError, ValueError):
                    continue
                out.add((etapa, curso, key))
                fam = competencias_materia_group_key(key) or key
                out.add((etapa, curso, fam))
            return out


def _latest_import_id() -> int | None:
    from db.enrolled_subjects import _latest_import_id as _lid

    return _lid()


def _materia_abrevs_for_key(*, materia_key: str, curso: int | None) -> list[str]:
    ensure_subject_catalog_schema()
    key = competencias_materia_group_key(materia_key) or (materia_key or "").strip()
    if not key:
        return []
    with get_db() as conn:
        with conn.cursor() as cur:
            if curso is not None:
                cur.execute(
                    """
                    SELECT materia_abrev, materia
                    FROM enrolled_subject_catalog
                    WHERE curso_asignatura = %s
                    """,
                    (int(curso),),
                )
            else:
                cur.execute(
                    """
                    SELECT materia_abrev, materia
                    FROM enrolled_subject_catalog
                    """
                )
            rows = cur.fetchall()
    out: list[str] = []
    for r in rows:
        mat = (r.get("materia") or "").strip()
        abrev = (r.get("materia_abrev") or "").strip()
        if not abrev:
            continue
        rk = competencias_materia_group_key(mat) or mat.casefold()
        if rk == key and abrev not in out:
            out.append(abrev)
    return out


def list_alumnos_evaluar(
    *,
    grupo: str,
    etapa: str,
    curso_asignatura: int | None,
    materia_key: str,
    pendiente: bool = False,
) -> list[str]:
    """Alumnos matriculados en la materia y grupo; si no hay, roster del grupo.

    Si ``pendiente`` es True, solo alumnos con la materia como pendiente
    (cursos anteriores); no se usa el roster completo del grupo.
    """
    ensure_enrolled_subjects_schema()
    nombre = (grupo or "").strip()
    if not nombre:
        return []

    import_id = _latest_import_id()
    key = competencias_materia_group_key(materia_key) or (materia_key or "").strip()
    alumnos: list[str] = []
    if pendiente:
        carac_sql = "TRIM(COALESCE(es.caracteristicas, '')) = %s"
    else:
        carac_sql = """
            (
              es.caracteristicas IS NULL
              OR TRIM(es.caracteristicas) = ''
              OR TRIM(es.caracteristicas) <> %s
            )
        """

    if import_id and key:
        abrevs = _materia_abrevs_for_key(
            materia_key=materia_key, curso=curso_asignatura
        )
        join_al = f"{sql_alumno_key('s.alumno')} = {sql_alumno_key('es.alumno')}"
        with get_db() as conn:
            with conn.cursor() as cur:
                if abrevs:
                    cur.execute(
                        f"""
                        SELECT DISTINCT TRIM(es.alumno) AS alumno
                        FROM enrolled_subjects es
                        LEFT JOIN students s
                          ON {join_al}
                        WHERE es.import_id = %s
                          AND TRIM(COALESCE(es.alumno, '')) <> ''
                          AND TRIM(es.materia_abrev) = ANY(%s)
                          AND {carac_sql}
                          AND LOWER(TRIM(COALESCE(NULLIF(TRIM(s.grupo), ''), es.nombre_grupo)))
                              = LOWER(TRIM(%s))
                        """,
                        (
                            import_id,
                            abrevs,
                            CARACTERISTICA_MATERIA_PENDIENTE,
                            nombre,
                        ),
                    )
                    alumnos = [
                        str(r["alumno"]).strip()
                        for r in cur.fetchall()
                        if r.get("alumno")
                    ]
                else:
                    cur.execute(
                        f"""
                        SELECT DISTINCT
                            TRIM(es.alumno) AS alumno,
                            TRIM(es.materia) AS materia,
                            c.materia AS catalog_materia
                        FROM enrolled_subjects es
                        LEFT JOIN enrolled_subject_catalog c
                          ON TRIM(c.materia_abrev) = TRIM(es.materia_abrev)
                        LEFT JOIN students s
                          ON {join_al}
                        WHERE es.import_id = %s
                          AND TRIM(COALESCE(es.alumno, '')) <> ''
                          AND {carac_sql}
                          AND LOWER(TRIM(COALESCE(NULLIF(TRIM(s.grupo), ''), es.nombre_grupo)))
                              = LOWER(TRIM(%s))
                        """,
                        (
                            import_id,
                            CARACTERISTICA_MATERIA_PENDIENTE,
                            nombre,
                        ),
                    )
                    seen: set[str] = set()
                    for r in cur.fetchall():
                        mat = (
                            r.get("catalog_materia") or r.get("materia") or ""
                        ).strip()
                        rk = competencias_materia_group_key(mat) or mat.casefold()
                        if rk != key:
                            continue
                        al = (r.get("alumno") or "").strip()
                        if al and al not in seen:
                            seen.add(al)
                            alumnos.append(al)

    if not alumnos:
        if pendiente:
            return []
        return get_students_by_group(nombre)
    alumnos.sort(key=normalize_for_sort)
    return alumnos


def list_notas_evaluar(
    *,
    etapa: str,
    curso_asignatura: int,
    materia_key: str,
    grupo: str,
    sesion: str | None = None,
) -> dict[tuple[str, str], Decimal]:
    """Mapa (alumno, criterio) -> nota."""
    ensure_competencias_evaluacion_schema()
    key = competencias_materia_group_key(materia_key) or (materia_key or "").strip()
    tbl = _table_notas(sesion)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT alumno, criterio, nota
                FROM {tbl}
                WHERE etapa = %s
                  AND curso_asignatura = %s
                  AND materia_key = %s
                  AND LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                """,
                (
                    (etapa or "").strip().lower(),
                    int(curso_asignatura),
                    key,
                    (grupo or "").strip(),
                ),
            )
            out: dict[tuple[str, str], Decimal] = {}
            for r in cur.fetchall():
                al = str(r["alumno"] or "").strip()
                cr = str(r["criterio"] or "").strip()
                if not al or not cr or r.get("nota") is None:
                    continue
                out[(al, cr)] = Decimal(str(r["nota"]))
            return out


def replace_notas_evaluar(
    *,
    etapa: str,
    curso_asignatura: int,
    materia_key: str,
    grupo: str,
    notas: dict[tuple[str, str], Decimal | None],
    updated_by: int | None = None,
    sesion: str | None = None,
    pendiente: bool = False,
) -> None:
    # Persiste nota_comp; solo rebuild descriptores si cambian calificaciones.
    ensure_competencias_evaluacion_schema()
    etapa_v = (etapa or "").strip().lower()
    curso = int(curso_asignatura)
    key = competencias_materia_group_key(materia_key) or (materia_key or "").strip()
    grupo_v = (grupo or "").strip()
    if not etapa_v or not key or not grupo_v:
        raise ValueError("Datos de evaluacion incompletos")

    prev = list_notas_evaluar(
        etapa=etapa_v,
        curso_asignatura=curso,
        materia_key=key,
        grupo=grupo_v,
        sesion=sesion,
    )
    nuevas: dict[tuple[str, str], Decimal] = {}
    for (alumno, criterio), nota in notas.items():
        al = (alumno or "").strip()
        cr = (criterio or "").strip()
        if not al or not cr or nota is None:
            continue
        nuevas[(al, cr)] = nota

    sin_cambios = prev == nuevas

    tbl = _table_notas(sesion)
    if not sin_cambios:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    DELETE FROM {tbl}
                    WHERE etapa = %s
                      AND curso_asignatura = %s
                      AND materia_key = %s
                      AND LOWER(TRIM(grupo)) = LOWER(TRIM(%s))
                    """,
                    (etapa_v, curso, key, grupo_v),
                )
                for (al, cr), nota in nuevas.items():
                    cur.execute(
                        f"""
                        INSERT INTO {tbl} (
                            etapa, curso_asignatura, materia_key, grupo,
                            alumno, criterio, nota, updated_at, updated_by
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), %s)
                        """,
                        (etapa_v, curso, key, grupo_v, al, cr, nota, updated_by),
                    )

    from db.competencias_materia_variables import contexto_ppd_phoras_materia

    ppd_map = contexto_ppd_phoras_materia(
        etapa=etapa_v,
        curso_asignatura=curso,
        materia_key=key,
        sesion=sesion,
        pendiente=pendiente,
    )["ppd_map"]
    por_alumno: dict[str, dict[str, Decimal | None]] = {}
    for (al, cr), nota in nuevas.items():
        por_alumno.setdefault(al, {})[cr] = nota
    notas_comp: dict[str, Decimal | None] = {
        al: compute_nota_comp(por_crit, ppd_map) for al, por_crit in por_alumno.items()
    }
    replace_notas_comp(
        etapa=etapa_v,
        curso_asignatura=curso,
        materia_key=key,
        grupo=grupo_v,
        notas=notas_comp,
        updated_by=updated_by,
        sesion=sesion,
    )

    if sin_cambios:
        return

    from db.competencias_alumno_descriptor import rebuild_alumno_descriptor_grupo
    from utils.group_stage import stage_of
    from db.groups import get_group_curso

    rebuild_alumno_descriptor_grupo(grupo_v, sesion=sesion)
    # Las competencias de extraordinaria reutilizan notas de ordinaria en materias
    # aprobadas: si cambian las ordinarias, hay que recalcular también la extra.
    if not _sesion_es_extraordinaria(sesion):
        curso_g = get_group_curso(grupo_v)
        if stage_of(grupo=grupo_v, curso=curso_g) == "bachillerato":
            rebuild_alumno_descriptor_grupo(grupo_v, sesion="extraordinaria")


NOTAS_ACTA_ESO: tuple[str, ...] = ("IN", "SU", "BI", "NT", "SB")
_NOTA_ACTA_ESO_A_NUM: dict[str, Decimal] = {
    "IN": Decimal(4),
    "SU": Decimal(5),
    "BI": Decimal(6),
    "NT": Decimal(7),
    "SB": Decimal(9),
}


def acta_es_cualitativa(etapa: str | None) -> bool:
    return (etapa or "").strip().lower() == "eso"


def codigo_nota_acta_eso(value: object) -> str:
    """Convierte nota numérica o código a IN/SU/BI/NT/SB."""
    if value is None:
        return ""
    text = str(value).strip().upper().replace("Í", "I")
    if not text:
        return ""
    aliases = {
        "INS": "IN",
        "INSUFICIENTE": "IN",
        "SUF": "SU",
        "SUFICIENTE": "SU",
        "BIEN": "BI",
        "NOTABLE": "NT",
        "SOB": "SB",
        "SOBRESALIENTE": "SB",
    }
    if text in NOTAS_ACTA_ESO:
        return text
    if text in aliases:
        return aliases[text]
    try:
        n = int(Decimal(text.replace(",", ".")).to_integral_value())
    except (InvalidOperation, ValueError, ArithmeticError):
        return ""
    if n <= 4:
        return "IN"
    if n == 5:
        return "SU"
    if n == 6:
        return "BI"
    if n <= 8:
        return "NT"
    return "SB"


def parse_nota_acta(raw: object, *, cualitativa: bool = False) -> Decimal | None:
    """nota_acta: en ESO código IN/SU/BI/NT/SB; en Bach entero 0-10. Vacío -> None."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    if cualitativa:
        code = codigo_nota_acta_eso(text)
        if code not in NOTAS_ACTA_ESO:
            raise ValueError("nota_acta de ESO debe ser IN, SU, BI, NT o SB")
        return _NOTA_ACTA_ESO_A_NUM[code]
    text = text.replace(",", ".")
    try:
        value = Decimal(text)
    except (InvalidOperation, ValueError):
        raise ValueError("nota_acta debe ser un entero entre 0 y 10") from None
    if value < 0 or value > 10 or value != value.to_integral_value():
        raise ValueError("nota_acta debe ser un entero entre 0 y 10")
    return value.to_integral_value()


def format_nota_acta_es(
    value: Decimal | float | int | str | None,
    *,
    cualitativa: bool = False,
) -> str:
    """nota_acta en pantalla: IN/SU/BI/NT/SB en ESO; entero en Bach."""
    if value is None or str(value).strip() == "":
        return ""
    if cualitativa:
        return codigo_nota_acta_eso(value)
    d = Decimal(str(value)).to_integral_value(rounding=ROUND_HALF_UP)
    return format(d, "f")


def parse_nota(raw: object) -> Decimal | None:
    """Calificacion por criterio: entre 0 y 10, hasta 2 decimales."""
    if raw is None:
        return None
    text = str(raw).strip().replace(",", ".")
    if not text:
        return None
    try:
        value = Decimal(text)
    except (InvalidOperation, ValueError):
        raise ValueError(
            "Solo se admiten notas entre 0 y 10 con hasta 2 decimales."
        ) from None
    if value < 0 or value > 10:
        raise ValueError(
            "Solo se admiten notas entre 0 y 10 con hasta 2 decimales."
        )
    if value != value.quantize(Decimal("0.01")):
        raise ValueError(
            "Solo se admiten notas entre 0 y 10 con hasta 2 decimales."
        )
    return value


def format_nota_es(value: Decimal | float | int | None) -> str:
    """Muestra la nota con coma decimal (p. ej. 7,5)."""
    if value is None:
        return ""
    d = Decimal(str(value))
    text = format(d, "f").rstrip("0").rstrip(".")
    if not text:
        text = "0"
    return text.replace(".", ",")


def format_nota_materia_es(value: Decimal | float | int | None) -> str:
    """Nota de materia en pantalla: un decimal redondeado (p. ej. 7,5 / 8,0)."""
    if value is None:
        return ""
    d = Decimal(str(value)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
    return format(d, "f").replace(".", ",")


def compute_nota_comp(
    notas_criterio: dict[str, Decimal | None],
    ppd_por_criterio: dict[str, Decimal],
) -> Decimal | None:
    """nota_comp: suma (calificacion * ppd) / 100.

    Si no hay PPD útil (vacío, no cruza con criterios calificados, o pesos 0),
    media aritmética de las notas de criterio presentes.
    """
    total = Decimal("0")
    weight = Decimal("0")
    for crit, ppd in (ppd_por_criterio or {}).items():
        nota = notas_criterio.get(crit)
        if nota is None:
            continue
        w = Decimal(str(ppd))
        weight += w
        total += Decimal(str(nota)) * w
    if weight > 0:
        return total / Decimal("100")
    vals = [Decimal(str(n)) for n in notas_criterio.values() if n is not None]
    if not vals:
        return None
    return sum(vals, Decimal("0")) / Decimal(len(vals))


compute_nota_materia = compute_nota_comp


def build_evaluacion_workbook(
    *,
    grupo: str,
    etapa: str,
    curso_asignatura: int,
    materia_key: str,
    materia_label: str,
    criterios: list[str],
    alumnos: list[str],
    notas: dict[tuple[str, str], Decimal] | None = None,
) -> openpyxl.Workbook:
    """Excel: fila 1 criterios, columna A alumnos; hoja _meta para validar la subida."""
    notas = notas or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = DATA_SHEET

    header_fill = PatternFill("solid", fgColor="CCFBF1")
    header_font = Font(bold=True)
    ws.cell(1, 1, "Alumno")
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).font = header_font
    for col, crit in enumerate(criterios, start=2):
        cell = ws.cell(1, col, crit)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_i, alumno in enumerate(alumnos, start=2):
        ws.cell(row_i, 1, alumno)
        for col, crit in enumerate(criterios, start=2):
            nota = notas.get((alumno, crit))
            if nota is not None:
                cell = ws.cell(row_i, col, format_nota_es(nota))
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "@"

    ws.column_dimensions["A"].width = 36
    for col in range(2, len(criterios) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 8

    meta = wb.create_sheet(META_SHEET)
    meta.append(["campo", "valor"])
    meta.append(["grupo", grupo])
    meta.append(["etapa", etapa])
    meta.append(["curso", int(curso_asignatura)])
    meta.append(["materia_key", materia_key])
    meta.append(["materia", materia_label])
    meta.append(["criterios", "|".join(criterios)])
    meta.sheet_state = "hidden"

    return wb


def build_evaluacion_xlsx_bytes(
    *,
    grupo: str,
    etapa: str,
    curso_asignatura: int,
    materia_key: str,
    materia_label: str,
    criterios: list[str],
    alumnos: list[str],
    notas: dict[tuple[str, str], Decimal] | None = None,
) -> bytes:
    """Plantilla .xlsx solo en RAM (sin openpyxl ni %TEMP%; evita bloqueos de AV)."""
    from utils.xlsx_export import evaluacion_plantilla_xlsx_bytes, two_sheets_xlsx_bytes

    if not notas:
        return evaluacion_plantilla_xlsx_bytes(
            grupo=grupo,
            etapa=etapa,
            curso_asignatura=curso_asignatura,
            materia_key=materia_key,
            materia_label=materia_label,
            criterios=criterios,
            alumnos=alumnos,
            data_sheet=DATA_SHEET,
            meta_sheet=META_SHEET,
        )

    data_rows: list[list[object]] = [["Alumno", *criterios]]
    for alumno in alumnos:
        row: list[object] = [alumno]
        for crit in criterios:
            nota = notas.get((alumno, crit))
            row.append(format_nota_es(nota) if nota is not None else "")
        data_rows.append(row)

    meta_rows: list[list[object]] = [
        ["campo", "valor"],
        ["grupo", grupo],
        ["etapa", etapa],
        ["curso", int(curso_asignatura)],
        ["materia_key", materia_key],
        ["materia", materia_label],
        ["criterios", "|".join(criterios)],
    ]
    return two_sheets_xlsx_bytes(
        sheet1_name=DATA_SHEET,
        sheet1_rows=data_rows,
        sheet2_name=META_SHEET,
        sheet2_rows=meta_rows,
        sheet2_hidden=True,
    )


def workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    """Guarda el workbook evitando %TEMP% del sistema (Avast borra openpyxl.*)."""
    import tempfile
    from pathlib import Path

    tmp_root = Path(__file__).resolve().parents[1] / ".cache" / "openpyxl_tmp"
    tmp_root.mkdir(parents=True, exist_ok=True)
    prev = tempfile.tempdir
    stream = io.BytesIO()
    try:
        tempfile.tempdir = str(tmp_root)
        wb.save(stream)
        return stream.getvalue()
    finally:
        tempfile.tempdir = prev
        for leftover in tmp_root.glob("openpyxl.*"):
            try:
                leftover.unlink(missing_ok=True)
            except OSError:
                pass


def parse_evaluacion_workbook(
    raw: bytes,
    *,
    expected_grupo: str,
    expected_etapa: str,
    expected_curso: int,
    expected_materia_key: str,
    expected_criterios: list[str],
) -> tuple[dict[tuple[str, str], Decimal | None] | None, str | None]:
    """
    Valida formato del Excel en blanco y extrae notas.
    Devuelve (mapa, error).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        return None, "No se pudo leer el archivo. Debe ser un Excel (.xlsx) valido."

    if META_SHEET not in wb.sheetnames or DATA_SHEET not in wb.sheetnames:
        return None, (
            "El archivo no tiene el formato esperado. "
            "Utilice unicamente la plantilla descargada con Descargar Excel en blanco."
        )

    meta = {str(r[0]).strip(): r[1] for r in wb[META_SHEET].iter_rows(min_row=2, values_only=True) if r and r[0]}
    key_exp = competencias_materia_group_key(expected_materia_key) or expected_materia_key
    key_file = competencias_materia_group_key(str(meta.get("materia_key") or "")) or str(
        meta.get("materia_key") or ""
    ).strip()

    if str(meta.get("grupo") or "").strip().casefold() != expected_grupo.strip().casefold():
        return None, "El Excel no corresponde a este grupo."
    if str(meta.get("etapa") or "").strip().lower() != expected_etapa.strip().lower():
        return None, "El Excel no corresponde a esta etapa."
    try:
        curso_file = int(meta.get("curso"))
    except (TypeError, ValueError):
        return None, "El Excel no indica un curso valido."
    if curso_file != int(expected_curso):
        return None, "El Excel no corresponde a este curso."
    if key_file != key_exp:
        return None, "El Excel no corresponde a esta materia."

    crit_file = [
        c.strip()
        for c in str(meta.get("criterios") or "").split("|")
        if c.strip()
    ]
    if crit_file != list(expected_criterios):
        return None, (
            "Los criterios del Excel no coinciden con los de esta materia. "
            "Descargue de nuevo la plantilla en blanco e intentelo otra vez."
        )

    ws = wb[DATA_SHEET]
    headers: list[str] = []
    for col in range(1, len(expected_criterios) + 2):
        val = ws.cell(1, col).value
        headers.append(str(val).strip() if val is not None else "")
    if not headers or headers[0].casefold() != "alumno":
        return None, (
            "La primera celda debe ser Alumno. "
            "No modifique la estructura de la plantilla descargada."
        )
    if headers[1:] != list(expected_criterios):
        return None, (
            "La fila de criterios no coincide con la plantilla. "
            "No reordene ni renombre las columnas de criterios."
        )

    notas: dict[tuple[str, str], Decimal | None] = {}
    row = 2
    while True:
        alumno_raw = ws.cell(row, 1).value
        if alumno_raw is None or str(alumno_raw).strip() == "":
            # Fin si no hay mas alumnos; permitir filas vacias intercaladas cortas
            empty_tail = True
            for peek in range(row, row + 3):
                if ws.cell(peek, 1).value not in (None, ""):
                    empty_tail = False
                    break
            if empty_tail:
                break
            row += 1
            continue
        alumno = str(alumno_raw).strip()
        for col, crit in enumerate(expected_criterios, start=2):
            raw_nota = ws.cell(row, col).value
            if raw_nota is None or str(raw_nota).strip() == "":
                notas[(alumno, crit)] = None
                continue
            try:
                nota = parse_nota(raw_nota)
            except ValueError:
                return None, (
                    f"Calificacion no valida para '{alumno}', criterio {crit}. "
                    "Use un numero entre 0 y 10 con hasta 2 decimales (p. ej. 7,5)."
                )
            if nota is None:
                notas[(alumno, crit)] = None
                continue
            notas[(alumno, crit)] = nota
        row += 1
        if row > 5000:
            break

    return notas, None


def safe_xlsx_filename(*parts: str) -> str:
    chunks = []
    for p in parts:
        t = re.sub(r"[^\w\-]+", "_", (p or "").strip(), flags=re.UNICODE)
        t = t.strip("_")[:40]
        if t:
            chunks.append(t)
    base = "_".join(chunks) or "evaluacion"
    return f"{base}.xlsx"


def criterios_codes(
    *,
    etapa: str,
    curso_asignatura: int,
    materia_key: str,
) -> list[str]:
    rows = list_criterios_materia(
        etapa=etapa,
        curso_asignatura=int(curso_asignatura),
        materia_key=materia_key,
    )
    return [str(r.get("criterio") or "").strip() for r in rows if r.get("criterio")]
